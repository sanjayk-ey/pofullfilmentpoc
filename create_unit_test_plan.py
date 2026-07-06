"""
create_unit_test_plan.py
Generates docs/Unit_Test_Plan_Interactive_Demo.xlsx — a business-friendly test
plan for the interactive, human-in-the-loop demo built around just TWO PO files:

    demo/Happy-Flow-PO.txt    -> the clean "everything works" flow
    demo/CSR-Approval-PO.txt  -> one PO that exercises every interactive CSR
                                 decision (unknown buyer, missing/wrong/obsolete
                                 SKU, zero qty, UOM conversion, partial ship-to,
                                 and the Pricing / Credit / Inventory / Logistics
                                 decision-layer gates)

The Pass/Fail column is filled by actually RUNNING the headless checks in
test_pipeline.py, so the numbers reflect real results.

Run:  python create_unit_test_plan.py
"""
import os
import io
import contextlib

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import test_pipeline as T

OUT_DIR = os.path.join(os.path.dirname(__file__), "docs")
os.makedirs(OUT_DIR, exist_ok=True)

NAVY = "1F2A44"; BLUE = "2563EB"; GREEN = "1E7D44"; RED = "B91C1C"; GREY = "F3F4F6"

# Business-friendly scenarios. `check` maps to the exact check name(s) recorded by
# test_pipeline so Pass/Fail is real.
ROWS = [
    ("T-01", "PO-1 happy path", "US-01 AC-01",
     "Clean PO with only the mandatory fields (PO#, PO date, buyer company + email, "
     "ship-to, delivery date, and line SKU/description/qty).",
     "Agent reads the PO, resolves the customer and buyer from the company name + "
     "email, and confirms all mandatory fields are present.",
     "PO card shows every field; no CSR decision needed.",
     ["happy: no missing mandatory fields", "happy: buyer_email is buyer (not vendor)"]),

    ("T-02", "PO-1 happy path", "US-02..US-12",
     "The clean PO runs end to end through all decision stages.",
     "Agent runs customer, buyer, product, compliance, pricing, budget, credit, "
     "inventory, logistics and execution — pausing for nothing.",
     "All stages green; order created and customer confirmation sent.",
     ["PO-1 happy path -> clean pass"]),

    ("T-03", "PO-2 exceptions", "US-01 AC-02",
     "Optional fields only mandatory ones required — a line gives description + qty "
     "with NO SKU.",
     "Agent does NOT reject; it identifies the product from the description against "
     "the catalog and asks the CSR to confirm / enter the SKU.",
     "CSR sees the recommended SKU with match confidence and Approve / Enter / "
     "Escalate options.",
     ["missing SKU identified by description", "PO-2 raised >=4 intake issues"]),

    ("T-04", "PO-2 exceptions", "US-04 (label independence)",
     "A line uses a non-catalog product code (customer's own label).",
     "Agent ignores the label and matches by description to the correct catalog SKU.",
     "CSR sees the identified SKU and approves it.",
     ["wrong SKU code identified by description"]),

    ("T-05", "PO-2 exceptions", "US-04 AC-04",
     "A line contains an obsolete / discontinued SKU.",
     "Agent finds the approved successor and shows compatibility, price impact and "
     "availability impact, and requires CSR approval before substituting.",
     "CSR sees the substitution recommendation and Approve / Reject / Escalate.",
     ["obsolete SKU -> substitute recommendation",
      "substitution has impacts (price/availability/compat)"]),

    ("T-06", "PO-1 happy path", "US-04 AC-02 (base UOM)",
     "PO omits UOM entirely on every line (as in the new demo PO format).",
     "Agent looks up the product's base UOM in Product Master and uses it — no "
     "CSR question needed for clear quantities. The Matched-products table "
     "hides the Converted / Conversion-logic columns when no numeric "
     "conversion is happening (so the happy-flow view stays clean); the "
     "inference is captured in the stage's audit trail.",
     "Matched-products table shows only SKU / Description / Family / "
     "Requested; audit trail records 'inferred from Product Master'.",
     ["missing UOM defaults to base UOM (no error)",
      "missing UOM audit trail records 'inferred from Product Master'",
      "no-conversion case hides Converted / Conversion logic columns"]),

    ("T-07", "PO-2 exceptions", "US-04 AC-02 (ambiguous qty)",
     "A line has a small qty for a product that ships in a case pack "
     "(e.g. qty=2 of a 24-pack cartridge).",
     "Agent recognises the ambiguity, presents 'individual pieces' vs "
     "'full packs' with the conversion maths (2 CASE = 48 EA), and asks CSR "
     "to confirm.",
     "CSR sees two explicit choices with converted-qty and conversion logic, "
     "plus Reject / Escalate.",
     ["UOM ambiguity detected for small qty vs pack size",
      "UOM ambiguity offers pack conversion 2 CASE = 48 EA"]),

    ("T-08", "internal", "US-04 AC-02 (approved rules)",
     "Approved conversion rules cover multiple product families "
     "(CARTRIDGE, VALVE, DRAIN, SEAL, FINISH) and universal (KG, DOZ, GR).",
     "Agent picks the family-specific rule when it exists and falls back to "
     "the universal ALL rule otherwise; shows original qty × factor = base qty.",
     "Pack conversions per family (CASE→EA×24 for CARTRIDGE, PALLET→EA×48 for "
     "VALVE, etc.) all round-trip correctly.",
     ["CASE->EA conversion produces 48 EA",
      "PALLET->EA conversion produces 48 EA (VALVE family)"]),

    ("T-09", "PO-2 exceptions", "US-02 AC-02",
     "Ship-to is given as a factory / location name only (partial address).",
     "Agent matches it to a registered ship-to for the customer and asks the CSR to "
     "confirm, or to type a different address.",
     "CSR sees the matched location (with ZIP) and Approve / Enter / Escalate.",
     ["partial ship-to name -> confirmation to 60639"]),

    ("T-10", "PO-2 exceptions", "US-11 AC-02",
     "After the CSR approves every recommendation, the order should complete.",
     "Agent resumes the pipeline from where it paused once each decision is made.",
     "All remaining stages pass; order created.",
     ["PO-2 after CSR approvals -> clean pass"]),

    ("T-11", "PO-1 happy path", "US-01 (master-data fallback)",
     "PO omits the optional Contact Person and Contract Reference header "
     "fields (customer sends only the mandatory ones). Delivery instructions "
     "belong to the specific PO transaction — the agent must NOT backfill "
     "them from master data.",
     "Agent backfills Contact Person from Buyer_Profiles (by email) and "
     "Contract Reference from the customer's ACTIVE Contracts row. Each "
     "backfilled value is tagged with a 'from master data' badge in the "
     "PO card and shown at 100% confidence (trusted source). Delivery "
     "instructions stay empty unless the PO itself carries them.",
     "PO card shows the two backfilled values with a blue chip and 100% "
     "confidence bars; Delivery Instructions shows 'not available' when "
     "the PO doesn't include them.",
     ["PO-1 contact person backfilled from Buyer_Profiles",
      "PO-1 contract reference backfilled from active Contracts",
      "PO-1 backfilled contact person shows 100% confidence",
      "PO-1 backfilled contract reference shows 100% confidence",
      "PO-1 delivery instructions NOT backfilled (per PO-only rule)"]),

    ("T-12", "PO-1 happy path", "US-06 (tax + shipping)",
     "The pricing engine must internally calculate BOTH tax and shipping and "
     "show them on-screen (customer PO does not include either).",
     "Agent looks up the Freight_Terms row for the customer (or the default) "
     "and computes base + per-KG shipping; then looks up the ship-to state "
     "in Tax_Rates and applies the correct sales-tax % on subtotal + "
     "surcharges. Both values appear in the 'Tax & shipping (AI-calculated)' "
     "table and roll up into the order total.",
     "'Order totals' shows: Subtotal, Surcharges, Freight/shipping, Sales "
     "tax (X% — State), Order total.",
     ["tax amount computed and > 0",
      "Illinois tax rate applied (~8.75%)",
      "freight amount computed and > 0",
      "pricing stage exposes 'Tax & shipping' table"]),

    ("T-13", "PO-1 & PO-2 (Excel)", "US-01 (Excel intake)",
     "Both demo POs are also delivered as .xlsx workbooks (real customers "
     "often send POs from Excel or as an Excel export from their ERP).",
     "Agent detects the workbook, parses the key-value header block plus the "
     "line-items table, and pushes the reconstructed PO through the same "
     "pipeline as the text version. Happy-Flow-PO completes end-to-end; "
     "CSR-Approval-PO raises the same interactive intake exceptions.",
     "CSR uploads the .xlsx and sees identical downstream behaviour.",
     ["Happy-Flow-PO.xlsx extracted",
      "Happy-Flow-PO.xlsx order lines = 3",
      "CSR-Approval-PO.xlsx extracted",
      "CSR-Approval-PO.xlsx order lines = 7",
      "Happy-Flow-PO.xlsx end-to-end clean pass"]),

    ("T-14", "PO with unknown buyer", "US-03 (interactive buyer resolution)",
     "The PO's buyer email is not in the Buyer Profiles master (typo / new "
     "employee not yet onboarded / customer-side email change).",
     "Agent detects that the email does not resolve, looks up the buyers "
     "already registered against this customer account, and presents them "
     "as a picker with 'Pick this buyer' buttons. CSR can also type a "
     "different buyer name / email or escalate before the authorization "
     "stage runs.",
     "CSR sees a buyer table (Buyer / Email / Role / Customer / Cost "
     "Center), individual 'Pick this buyer' buttons, a free-text entry "
     "field, plus Reject / Escalate.",
     ["unknown buyer email raises UNRESOLVED_BUYER",
      "UNRESOLVED_BUYER lists buyers for the customer"]),

    ("T-15", "PO with zero quantity", "US-01 (interactive qty correction)",
     "One order line has quantity = 0 (customer left the field blank in "
     "their form / accidental clear).",
     "Agent does NOT hard-block intake. It raises INVALID_QUANTITY for that "
     "specific line, showing the SKU and the current (invalid) quantity, "
     "and asks the CSR to type the correct positive quantity before the "
     "pipeline resumes.",
     "CSR sees a small line-context table plus a 'Type the correct quantity' "
     "field with Use my entry / Reject / Escalate.",
     ["zero-qty line does NOT hard-block intake",
      "zero quantity raises INVALID_QUANTITY intake issue",
      "INVALID_QUANTITY targets the correct line"]),

    ("T-16", "PO-2 exceptions", "US-04 (manual override on substitution)",
     "The AI recommends an approved substitute for an obsolete SKU, but the "
     "CSR wants to force a completely different SKU instead.",
     "Substitution card now offers a 'Use my entry' text field alongside "
     "Approve. Whatever the CSR types replaces the substitute; the audit "
     "trail records the manual override.",
     "CSR sees the substitution card with the extra text field and the "
     "'✍️ Use my entry' button.",
     ["SUBSTITUTE_SKU includes 'enter' manual-entry action"]),
]

HEADERS = ["Test ID", "PO file", "AC", "Scenario (plain words)", "What the agent does",
           "What the CSR sees / does", "Result (auto)", "Pass / Fail (auto)", "Tester sign-off"]
WIDTHS = [8, 20, 22, 40, 44, 40, 16, 14, 16]


def _border():
    s = Side(style="thin", color="C7CDD6")
    return Border(left=s, right=s, top=s, bottom=s)


def banner(ws, text, ncols, height=28, fill=NAVY, size=14):
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ncols)
    c = ws.cell(row=ws.max_row, column=1, value=text)
    c.font = Font(bold=True, size=size, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[ws.max_row].height = height


def run_checks():
    """Run the headless checks and return {check_name: ok}."""
    T._results.clear()
    with contextlib.redirect_stdout(io.StringIO()):
        from modules import duplicate_checker as dup
        dup.reset_store()
        T.test_extraction()
        T.test_intake_resolution()
        T.test_uom_conversion()
        T.test_end_to_end()
        T.test_master_data_backfill()
        T.test_tax_and_shipping()
        T.test_excel_pos()
        T.test_buyer_and_quantity_issues()
    return {name: ok for name, ok, _ in T._results}


def main():
    results = run_checks()
    total_checks = len(results)
    passed_checks = sum(1 for v in results.values() if v)

    row_status = []
    for r in ROWS:
        checks = r[6]
        ok = all(results.get(c, False) for c in checks)
        row_status.append(ok)
    n_pass = sum(row_status)

    wb = openpyxl.Workbook()

    # ── Tab 1: How To Test ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "How To Test"
    ws.sheet_view.showGridLines = False
    ws.append([""]); banner(ws, "PO FULFILMENT AI — INTERACTIVE DEMO TEST PLAN", 2, 30, NAVY, 16)
    intro = [
        "",
        f"LATEST AUTOMATED RUN:  {passed_checks} of {total_checks} underlying checks PASSED  "
        f"({n_pass} of {len(ROWS)} business test cases green).",
        "",
        "THE DEMO USES ONLY TWO PO FILES",
        "  • demo/Happy-Flow-PO.txt    — the clean 'everything works' order.",
        "  • demo/CSR-Approval-PO.txt  — one order that triggers every CSR decision.",
        "",
        "HOW TO SEND A PO TO THE AGENT",
        "  Open the .txt file, copy all the text, paste it into the chat box, and press Enter.",
        "",
        "WHAT MAKES THIS DEMO DIFFERENT",
        "  The agent no longer just stops on a problem. When something is missing, wrong,",
        "  obsolete, partial, or needs conversion, it THINKS out loud, checks the master data,",
        "  proposes the best fix, and asks the CSR to Approve, Reject, or Escalate — or to type",
        "  a correction. Approve resumes the order; Reject stops it; Escalate routes it to the",
        "  right team.",
        "",
        "ONLY THESE FIELDS ARE MANDATORY (everything else is optional / looked up):",
        "  PO Number · PO Date · Buyer Company + Buyer Email · Ship-To (full, partial or name)",
        "  · Requested Delivery Date · per line: SKU, Description, Quantity  (UOM is optional).",
        "",
        "HOW TO RUN EACH TEST",
        "  1. Go to the 'Test Cases' tab.  2. Load the listed PO file.  3. Follow the scenario.",
        "  4. Compare what you see to 'What the CSR sees / does'.  5. Sign off in the last column.",
    ]
    for line in intro:
        ws.append([line]); c = ws.cell(row=ws.max_row, column=1)
        if line.startswith("LATEST AUTOMATED"):
            c.font = Font(bold=True, size=12, color=GREEN if passed_checks == total_checks else RED)
        elif line.strip().isupper() or line.strip().startswith("ONLY THESE"):
            c.font = Font(bold=True, size=11, color=BLUE)
        else:
            c.font = Font(size=11)
    ws.column_dimensions["A"].width = 108

    # ── Tab 2: Test Cases ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Test Cases")
    ws2.sheet_view.showGridLines = False
    ws2.append([""])
    banner(ws2, f"TEST CASES  —  {n_pass}/{len(ROWS)} green in the latest automated run",
            len(HEADERS), 28, NAVY, 14)
    ws2.append(HEADERS); hr = ws2.max_row
    for ci, h in enumerate(HEADERS, 1):
        cell = ws2.cell(row=hr, column=ci, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws2.row_dimensions[hr].height = 32

    palette = {"PO-1 happy path": "EAF1FB", "PO-2 exceptions": "FFF7ED"}
    for r, ok in zip(ROWS, row_status):
        tid, pofile, ac, scenario, agent, csr, checks = r
        ws2.append([tid, pofile, ac, scenario, agent, csr,
                    "PASS" if ok else "FAIL", "PASS" if ok else "FAIL", ""])
        ri = ws2.max_row
        fill = palette.get(pofile, "FFFFFF")
        for ci in range(1, len(HEADERS) + 1):
            cell = ws2.cell(row=ri, column=ci)
            cell.font = Font(size=10)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _border()
        ws2.cell(row=ri, column=1).font = Font(size=10, bold=True)
        for col in (7, 8):
            pf = ws2.cell(row=ri, column=col)
            pf.font = Font(size=10, bold=True, color="FFFFFF")
            pf.fill = PatternFill("solid", fgColor=GREEN if ok else RED)
            pf.alignment = Alignment(horizontal="center", vertical="center")
    for ci, w in enumerate(WIDTHS, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A4"

    out = os.path.join(OUT_DIR, "Unit_Test_Plan_Interactive_Demo.xlsx")
    wb.save(out)
    print(f"Created: {out}  ({len(ROWS)} business cases, {n_pass} green; "
          f"{passed_checks}/{total_checks} underlying checks passed)")


if __name__ == "__main__":
    main()
