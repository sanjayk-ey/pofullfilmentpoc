"""
Generates the customer master data workbook: mock-data/customer-master-data.xlsx
Run once:  python create_customer_master_data_excel.py

Sheets:
  Customer_Master   - customer accounts + ERP customer records + CRM/account records
  Account_Hierarchy - branch -> regional division -> global parent mapping
  Ship_To_Master    - ship-to locations with ZIP and owning branch
  Hierarchy_Rules   - rules defined at each hierarchy level (most specific wins)

NOTE: For the POC there is NO real ERP / CRM / WMS / OMS / SMTP. This workbook
simulates the customer-side "Systems / Data Needed" for account validation.
Other master data (e.g. product master) will live in its own workbook.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

OUT_DIR = os.path.join(os.path.dirname(__file__), "mock-data")
os.makedirs(OUT_DIR, exist_ok=True)

NAVY  = "1E3A5F"
LBLUE = "DBEAFE"
thin  = Side(style="thin", color="AAAAAA")


def _border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def write_sheet(ws, title, headers, rows, widths):
    # Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Header row
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
    ws.row_dimensions[2].height = 22

    # Data rows
    for ri, rowdata in enumerate(rows, start=3):
        for ci, val in enumerate(rowdata, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=11)
            cell.fill = PatternFill("solid", fgColor=LBLUE if ri % 2 == 0 else "FFFFFF")
            cell.alignment = Alignment(vertical="center")
            cell.border = _border()

    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w


wb = openpyxl.Workbook()

# ── Sheet 1: Customer_Master (incl. ERP + CRM records) ─────────────────────────
ws1 = wb.active
ws1.title = "Customer_Master"
write_sheet(
    ws1,
    "CUSTOMER MASTER  (includes ERP customer records + CRM / account records)",
    ["customer_account", "company_name", "status", "branch_id",
     "erp_customer_id", "crm_account_id", "parent_account_id", "hierarchy_level",
     "customer_tier", "payment_terms", "customer_class", "distributor_authorization",
     "default_currency"],
    [
        # customer_class identifies the type of trade partner (Distributor /
        # Contractor / Retailer) and distributor_authorization records whether
        # the account is an authorized distributor for the brand. These, together
        # with customer_tier and payment_terms, drive the Customer Validation
        # decision layer and are shown in the Resolved Account Hierarchy.
        ["CUST-1001", "Great Lakes Plumbing Supply Co",  "ACTIVE", "BR-GLP-MW", "ERP-GL-7781", "CRM-GLP-001", "GP-CONT", "branch", "Strategic", "NET30", "Distributor", "Authorized Distributor", "USD"],
        ["CUST-1002", "Eastern Kitchen & Bath Distributors", "ACTIVE", "BR-GLP-NE", "ERP-GL-7782", "CRM-GLP-002", "GP-CONT", "branch", "Preferred", "NET45", "Distributor", "Authorized Distributor", "USD"],
        ["CUST-2001", "Continental Canada Distribution", "ACTIVE", "BR-GLP-CA", "ERP-GL-7790", "CRM-GLP-010", "GP-CONT", "branch", "Preferred", "NET30", "Distributor", "Authorized Distributor", "CAD"],
        ["CUST-5001", "Pacific Coast Bath & Kitchen",    "ACTIVE", "BR-PCBK-WEST", "ERP-PC-3310", "CRM-PCBK-001", "GP-WSH", "branch", "Standard", "NET60", "Retailer", "Non-Distributor", "USD"],
        ["CUST-7000", "Midtown Building Supply",         "ACTIVE", "BR-GLP-MW", "ERP-MT-9001", "CRM-MT-001", "GP-CONT", "branch", "Standard", "PREPAYMENT", "Contractor", "Non-Distributor", "USD"],
        ["CUST-7000", "Midtown Building Supply (Legacy)","ACTIVE", "BR-GLP-NE", "ERP-MT-9002", "CRM-MT-002", "GP-CONT", "branch", "Standard", "PREPAYMENT", "Contractor", "Non-Distributor", "USD"],
    ],
    [18, 32, 10, 16, 16, 18, 16, 14, 12, 14, 16, 22, 14],
)

# ── Sheets 1b + 1c: Order_History + Order_History_Lines ────────────────────────
# Buying history is modelled the same way as the reference Sample_data.xlsx: as
# transactional PAST ORDERS (Order_History = header) and their line items
# (Order_History_Lines = SKU/qty/price detail). There is intentionally NO
# pre-summarised "buying history" table — the Customer Validation and Product
# Match decision layers derive the summary (tenure, order count, lifetime value,
# average order value, frequently ordered families / SKUs, last order) directly
# from these transactions at runtime, exactly like a real ERP order history.
#
# Product catalogue used to build realistic historical lines (mirrors the
# product master: sku -> (family, base_uom, list_price)).
_PRODUCTS = {
    "SKU-CTG-4520": ("CARTRIDGE",  "EA",  12.50),
    "SKU-DRN-3010": ("DRAIN",      "EA",  45.00),
    "SKU-VLV-2201": ("VALVE",      "EA",  88.00),
    "SKU-SHS-7700": ("SHOWERSYS",  "EA", 1450.00),
    "SKU-FIN-9100": ("FINISH",     "GAL",  9.75),
    "SKU-FIN-9200": ("FINISH",     "GAL", 14.00),
    "SKU-SEL-1150": ("SEAL",       "EA",   6.25),
}

# Contract-style discount applied to historical unit prices by customer tier so
# that lifetime value / average order value look realistic per account.
_TIER_DISCOUNT = {"Strategic": 0.12, "Preferred": 0.08, "Standard": 0.03}

# Per-customer purchase history:
#   customer_account -> (tier, currency, [ (po_number, order_date, order_status,
#                                           [ (sku, qty), ... ]) ])
_ORDER_HISTORY = {
    "CUST-1001": ("Strategic", "USD", [
        ("USPO-24-3110", "2024-09-15", "auto-approved", [("SKU-CTG-4520", 400), ("SKU-SEL-1150", 300)]),
        ("USPO-25-3204", "2025-01-20", "auto-approved", [("SKU-CTG-4520", 500), ("SKU-VLV-2201", 60)]),
        ("USPO-25-3388", "2025-05-12", "approved",      [("SKU-CTG-4520", 350), ("SKU-SEL-1150", 250)]),
        ("USPO-25-3512", "2025-09-08", "auto-approved", [("SKU-VLV-2201", 80),  ("SKU-CTG-4520", 300)]),
        ("USPO-26-3640", "2026-02-18", "auto-approved", [("SKU-CTG-4520", 600), ("SKU-SEL-1150", 400)]),
        ("USPO-26-3781", "2026-06-20", "auto-approved", [("SKU-CTG-4520", 450), ("SKU-VLV-2201", 40)]),
    ]),
    "CUST-1002": ("Preferred", "USD", [
        ("USPO-24-4102", "2024-11-10", "auto-approved", [("SKU-DRN-3010", 120), ("SKU-SEL-1150", 200)]),
        ("USPO-25-4231", "2025-04-22", "approved",      [("SKU-DRN-3010", 150)]),
        ("USPO-25-4390", "2025-10-05", "auto-approved", [("SKU-DRN-3010", 100), ("SKU-SEL-1150", 180)]),
        ("USPO-26-4488", "2026-03-14", "auto-approved", [("SKU-DRN-3010", 130)]),
        ("USPO-26-4602", "2026-05-28", "auto-approved", [("SKU-DRN-3010", 90),  ("SKU-SEL-1150", 160)]),
    ]),
    "CUST-2001": ("Preferred", "CAD", [
        ("CAPO-24-5101", "2024-12-01", "auto-approved", [("SKU-VLV-2201", 60),  ("SKU-CTG-4520", 200)]),
        ("CAPO-25-5230", "2025-06-15", "approved",      [("SKU-VLV-2201", 75)]),
        ("CAPO-25-5377", "2025-11-20", "auto-approved", [("SKU-VLV-2201", 50),  ("SKU-CTG-4520", 150)]),
        ("CAPO-26-5490", "2026-06-01", "auto-approved", [("SKU-VLV-2201", 65)]),
    ]),
    "CUST-5001": ("Standard", "USD", [
        ("USPO-25-6110", "2025-03-10", "approved",      [("SKU-FIN-9100", 40),  ("SKU-FIN-9200", 25)]),
        ("USPO-25-6244", "2025-08-18", "auto-approved", [("SKU-SHS-7700", 4)]),
        ("USPO-26-6351", "2026-01-25", "auto-approved", [("SKU-FIN-9100", 60)]),
        ("USPO-26-6470", "2026-04-15", "approved",      [("SKU-SHS-7700", 3),   ("SKU-FIN-9200", 30)]),
    ]),
    "CUST-7000": ("Standard", "USD", [
        ("USPO-24-7101", "2024-07-12", "approved",      [("SKU-DRN-3010", 100), ("SKU-CTG-4520", 150)]),
        ("USPO-25-7230", "2025-02-08", "approved",      [("SKU-DRN-3010", 80)]),
        ("USPO-25-7388", "2025-08-30", "on-hold",       [("SKU-CTG-4520", 120), ("SKU-DRN-3010", 90)]),
        ("USPO-26-7455", "2026-03-10", "on-hold",       [("SKU-DRN-3010", 70)]),
    ]),
}


def _build_order_history():
    """Expand _ORDER_HISTORY into header rows and line rows with computed totals."""
    header_rows = []
    line_rows = []
    for acct, (tier, currency, orders) in _ORDER_HISTORY.items():
        disc = _TIER_DISCOUNT.get(tier, 0.0)
        for oi, (po_number, order_date, status, lines) in enumerate(orders, 1):
            order_id = f"ORD-{acct.split('-')[-1]}-{oi:03d}"
            order_total = 0.0
            for li, (sku, qty) in enumerate(lines, 1):
                family, uom, list_price = _PRODUCTS[sku]
                unit_price = round(list_price * (1 - disc), 2)
                line_total = round(qty * unit_price, 2)
                order_total += line_total
                line_rows.append([
                    f"OL-{order_id}-{li}", order_id, sku, family,
                    qty, uom, unit_price, line_total,
                ])
            header_rows.append([
                order_id, acct, po_number, order_date, status,
                round(order_total, 2), currency,
            ])
    return header_rows, line_rows


_oh_headers, _oh_lines = _build_order_history()

ws1b = wb.create_sheet("Order_History")
write_sheet(
    ws1b,
    "ORDER HISTORY  (past orders — header; source for buying-history derivation)",
    ["order_id", "customer_account", "po_number", "order_date",
     "order_status", "order_total", "currency"],
    _oh_headers,
    [18, 18, 16, 14, 16, 14, 10],
)

ws1c = wb.create_sheet("Order_History_Lines")
write_sheet(
    ws1c,
    "ORDER HISTORY LINES  (past order line items — SKU / qty / price detail)",
    ["order_line_id", "order_id", "sku", "product_family",
     "quantity", "uom", "unit_price", "line_total"],
    _oh_lines,
    [22, 18, 16, 16, 12, 10, 12, 14],
)

# ── Sheet 2: Account_Hierarchy ─────────────────────────────────────────────────
ws2 = wb.create_sheet("Account_Hierarchy")
write_sheet(
    ws2,
    "ACCOUNT HIERARCHY  (branch -> regional division -> global parent)",
    ["branch_id", "branch_name", "regional_division_id", "regional_division_name",
     "global_parent_id", "global_parent_name"],
    [
        ["BR-GLP-MW", "Great Lakes Midwest Branch",   "RD-CONT-NA", "Continental North America", "GP-CONT", "Continental Building Products Group"],
        ["BR-GLP-NE", "Great Lakes Northeast Branch", "RD-CONT-NA", "Continental North America", "GP-CONT", "Continental Building Products Group"],
        ["BR-GLP-CA", "Great Lakes Canada Branch",    "RD-CONT-CA", "Continental Canada",        "GP-CONT", "Continental Building Products Group"],
        ["BR-PCBK-WEST", "Pacific Coast West Branch",  "RD-WSH-NA", "Western Supply North America", "GP-WSH", "Western Supply Holdings"],
    ],
    [16, 26, 20, 24, 16, 24],
)

# ── Sheet 3: Ship_To_Master ────────────────────────────────────────────────────
ws3 = wb.create_sheet("Ship_To_Master")
write_sheet(
    ws3,
    "SHIP-TO MASTER  (ship-to locations matched by ZIP)",
    ["ship_to_id", "name", "address", "zip", "branch_id", "status",
     "customer_account", "city", "state", "country", "is_primary",
     "preferred_warehouse_id", "split_shipment_allowed", "backorder_tolerance_days",
     "default_delivery_instructions"],
    [
        ["ST-CHI-001", "Great Lakes Plumbing - Chicago DC", "4500 West Diversey Avenue, Chicago, IL", "60639", "BR-GLP-MW", "ACTIVE", "CUST-1001", "Chicago", "IL", "US", "Y", "DC-CHI-01", "N", 0,
         "Deliver Mon-Fri 8am-4pm at DC receiving dock 3. Notify John Miller (312-555-0140) 30 min before arrival."],
        ["ST-DET-002", "Great Lakes Plumbing - Detroit Branch", "1200 Woodward Avenue, Detroit, MI", "48201", "BR-GLP-MW", "ACTIVE", "CUST-1001", "Detroit", "MI", "US", "N", "DC-DET-02", "Y", 2,
         "Deliver Tue-Fri 9am-3pm at branch dock B. Call 313-555-0210 upon arrival."],
        ["ST-NYC-003", "Eastern Kitchen & Bath - New York DC", "55 Water Street, New York, NY", "10001", "BR-GLP-NE", "ACTIVE", "CUST-1002", "New York", "NY", "US", "Y", "DC-CHI-01", "Y", 3,
         "Delivery via freight elevator (rear entrance). Weekdays 7am-2pm only."],
        ["ST-LON-004", "Continental Canada - Toronto Depot", "120 Bremner Boulevard, Toronto, ON", "M5J2N1", "BR-GLP-CA", "ACTIVE", "CUST-2001", "Toronto", "ON", "CA", "Y", "DC-DET-02", "Y", 5,
         "Commercial delivery only. Provide customs documents in advance."],
        ["ST-LA-005", "Pacific Coast - Los Angeles DC", "800 South Hope Street, Los Angeles, CA", "90001", "BR-PCBK-WEST", "ACTIVE", "CUST-5001", "Los Angeles", "CA", "US", "Y", "DC-LA-05", "Y", 2,
         "Deliver 7am-3pm. Curbside drop only; no forklift on-site."],
        ["ST-AK-006", "Great Lakes - Ketchikan Project Site", "1 Industrial Rd, Ketchikan, AK", "99950", "BR-GLP-MW", "ACTIVE", "CUST-1001", "Ketchikan", "AK", "US", "N", "DC-LA-05", "Y", 7,
         "Remote project site — coordinate with site supervisor 48 hours before arrival."],
        ["ST-CA-007", "Great Lakes - Beverly Hills Showroom Project", "9000 Sunset Blvd, Beverly Hills, CA", "90210", "BR-GLP-MW", "ACTIVE", "CUST-1001", "Beverly Hills", "CA", "US", "N", "DC-LA-05", "N", 0,
         "White-glove delivery required. Do NOT leave unattended."],
        ["ST-CA-008", "Great Lakes - Malibu Coast Project",             "22200 Pacific Coast Hwy, Malibu, CA", "90265", "BR-GLP-MW", "ACTIVE", "CUST-1001", "Malibu",        "CA", "US", "N", "DC-LA-05", "Y", 3,
         "Project site — access via service road. Contact foreman at 310-555-0090."],
    ],
    [14, 32, 40, 10, 16, 10, 16, 14, 8, 9, 10, 20, 18, 20, 60],
)

# ── Sheet 4: Hierarchy_Rules ───────────────────────────────────────────────────
# The `fulfillment_rule` column stores a RULE PROFILE ID. The actual structured
# business rules (preferred warehouse, split allowed, backorder allowed, etc.)
# are defined in the Fulfillment_Rules sheet below and looked up by this ID.
ws4 = wb.create_sheet("Hierarchy_Rules")
write_sheet(
    ws4,
    "HIERARCHY RULES  (blank = rule not set at that level; most specific level wins)",
    ["level_type", "level_id", "pricing_tier", "product_visibility",
     "budget_limit", "approval_routing", "fulfillment_rule"],
    [
        # Global parents
        ["global_parent", "GP-CONT", "TIER-3", "GLOBAL_CATALOG",  5000000, "CORPORATE_PROCUREMENT", "RULE-GLOBAL-STD"],
        ["global_parent", "GP-WSH",  "TIER-1", "STANDARD_CATALOG", 2000000, "CORPORATE_PROCUREMENT", "RULE-GLOBAL-STD"],
        # Regional divisions
        ["regional_division", "RD-CONT-NA", "TIER-2",    "", "", "REGIONAL_MANAGER",    "RULE-NA-STD"],
        ["regional_division", "RD-CONT-CA", "TIER-2-CA", "", "", "CA_REGIONAL_MANAGER", "RULE-CA-STD"],
        ["regional_division", "RD-WSH-NA", "TIER-1",  "", "", "REGIONAL_MANAGER",    ""],
        # Branches
        ["branch", "BR-GLP-MW", "", "", 500000, "BRANCH_MANAGER", "RULE-MW-STD"],
        ["branch", "BR-GLP-NE", "", "", 350000, "BRANCH_MANAGER", ""],
        ["branch", "BR-GLP-CA", "", "", 300000, "BRANCH_MANAGER", ""],
        ["branch", "BR-PCBK-WEST", "", "", 400000, "BRANCH_MANAGER", "RULE-WEST-STD"],
        # Ship-to (most specific level — overrides any branch/regional default)
        ["ship_to", "ST-CHI-001", "", "", "", "SITE_SUPERVISOR", "RULE-CHI-PRIORITY"],
        ["ship_to", "ST-DET-002", "", "", "", "", "RULE-DET-NO-SPLIT"],
        ["ship_to", "ST-NYC-003", "", "", "", "", "RULE-NYC-BACKORDER-OK"],
        ["ship_to", "ST-LON-004", "", "", "", "", "RULE-CA-STD"],
        ["ship_to", "ST-LA-005", "", "", "", "", "RULE-LA-RESTRICTED"],
        ["ship_to", "ST-CA-007", "", "", "", "", "RULE-LARGE-MOQ"],
        ["ship_to", "ST-CA-008", "", "", "", "", "RULE-COAST-TO-COAST"],
    ],
    [18, 16, 12, 18, 14, 22, 22],
)

# ── Sheet 5: Fulfillment_Rules ─────────────────────────────────────────────────
# Each row defines the actual business rules behind a fulfillment_rule profile
# referenced from Hierarchy_Rules. Validators in US-09 (inventory) and US-10
# (logistics) read these columns to decide:
#   - preferred_warehouse        first DC to attempt fulfillment from
#   - alternate_warehouses       fallback DCs to try (in order)
#   - restricted_warehouses      DCs that must NEVER be used for this customer
#   - split_shipment_allowed     Y = order may be split across DCs
#                                N = must ship in a single shipment
#   - backorder_allowed          Y = partial fulfillment OK; remainder backordered
#                                N = if not fully available, raise an exception
#   - max_backorder_days         max days the customer will accept for backorder
#   - min_order_qty              minimum total order quantity (lines aggregated)
#   - delivery_sla_days          target days from order to delivery
#   - allocation_priority        used when inventory is constrained
ws5 = wb.create_sheet("Fulfillment_Rules")
write_sheet(
    ws5,
    "FULFILLMENT RULES  (business rules referenced by hierarchy fulfillment_rule)",
    ["rule_id", "rule_name", "preferred_warehouse", "alternate_warehouses",
     "restricted_warehouses", "split_shipment_allowed", "backorder_allowed",
     "max_backorder_days", "min_order_qty", "delivery_sla_days",
     "allocation_priority", "description"],
    [
        ["RULE-GLOBAL-STD",       "Global standard",
         "DC-CHI-01", "DC-DET-02,DC-LA-05", "",
         "Y", "Y", 15, 1, 7, "SILVER",
         "Default global rule. Ship from nearest DC; split and backorder allowed."],

        ["RULE-NA-STD",           "North America standard",
         "DC-CHI-01", "DC-DET-02", "",
         "Y", "Y", 10, 1, 5, "SILVER",
         "Ship from Chicago first, then Detroit. Split and backorder allowed."],

        ["RULE-CA-STD",           "Canada standard",
         "DC-TOR-10", "", "",
         "Y", "N", 0, 1, 7, "SILVER",
         "Ship from Toronto depot only. Split allowed; no backorder (must be fully available)."],

        ["RULE-MW-STD",           "Midwest standard",
         "DC-CHI-01", "DC-DET-02", "",
         "Y", "Y", 10, 1, 4, "SILVER",
         "Ship from Chicago (preferred) or Detroit; split and short backorder allowed."],

        ["RULE-WEST-STD",         "West Coast standard",
         "DC-LA-05", "", "",
         "Y", "Y", 7, 1, 4, "SILVER",
         "Ship from LA DC. Split allowed; short backorder window."],

        ["RULE-CHI-PRIORITY",     "Priority Chicago (GOLD)",
         "DC-CHI-01", "DC-DET-02", "",
         "Y", "Y", 5, 1, 2, "GOLD",
         "Ship priority from Chicago DC. Same-day priority; minor backorder allowed."],

        ["RULE-DET-NO-SPLIT",     "Detroit single-shipment",
         "DC-DET-02", "DC-CHI-01", "",
         "N", "N", 0, 1, 2, "SILVER",
         "Customer requires one shipment, no split, no backorder, within a tight "
         "2-day delivery SLA. Full quantity must come from a single DC or the order "
         "is held for CSR review."],

        ["RULE-NYC-BACKORDER-OK", "NYC backorder allowed",
         "DC-NYC-03", "DC-CHI-01", "",
         "Y", "Y", 20, 1, 6, "SILVER",
         "Customer accepts long backorder windows. Ship what's available from NYC, "
         "backorder the rest up to 20 days."],

        ["RULE-LA-RESTRICTED",    "LA DC restricted",
         "DC-CHI-01", "DC-DET-02", "DC-LA-05",
         "Y", "Y", 10, 1, 6, "SILVER",
         "LA DC is restricted (customer audit failure). Must ship from Chicago/Detroit "
         "even if LA has stock."],

        ["RULE-LARGE-MOQ",        "Large minimum order quantity",
         "DC-CHI-01", "DC-DET-02", "",
         "Y", "Y", 7, 500, 5, "GOLD",
         "Bulk customer. Minimum total order quantity = 500 units. Smaller orders "
         "are rejected to protect margins."],

        ["RULE-COAST-TO-COAST",   "Coast-to-coast multi-DC optimization",
         "DC-CHI-01", "DC-LA-05,DC-DET-02", "",
         "Y", "Y", 10, 1, 7, "GOLD",
         "Project sites reachable from Chicago (preferred), Los Angeles, or Detroit. "
         "Optimization stage evaluates all three, splits allowed. Winner is chosen "
         "by lowest freight cost among feasible plans meeting the 7-day SLA."],
    ],
    [22, 28, 18, 22, 22, 14, 14, 14, 14, 14, 14, 48],
)

out = os.path.join(OUT_DIR, "customer-master-data.xlsx")
wb.save(out)
print("Created:", out)
