"""
create_demo_excel_pos.py
Generates the two demo POs in .xlsx format alongside their .txt counterparts:

  demo/Happy-Flow-PO.xlsx    — clean happy-path PO (straight-through auto decision)
  demo/CSR-Approval-PO.xlsx  — one PO that walks every interactive CSR approval
                               gate (intake + Pricing, Credit, Inventory, Logistics)

The layout is intentionally the same "standard US PO" shape a customer would
send:
    * two-column header block (label | value)
    * blank row separator
    * order-lines table (Item # | Product Code | Description | Qty)

`modules.excel_parser.parse_excel` reads this shape and hands the reconstructed
text to the same POExtractor the plain-text path uses, so both formats end up
in the same downstream pipeline.

Run once whenever the demo POs change:
    python create_demo_excel_pos.py
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.join(os.path.dirname(__file__), "demo")
os.makedirs(OUT_DIR, exist_ok=True)

NAVY = "1E3A5F"
BLUE = "2563EB"
LBLUE = "DBEAFE"
thin = Side(style="thin", color="AAAAAA")


def _border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_po(path, po_number, po_date, company, email, ship_to_lines,
              requested_delivery_date, line_items, title="PURCHASE ORDER"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO"

    # ── Banner ───────────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Header key/value block (columns A,B) ─────────────────────────────────
    def put(row, label, value):
        a = ws.cell(row=row, column=1, value=label)
        a.font = Font(bold=True)
        a.fill = PatternFill("solid", fgColor=LBLUE)
        a.alignment = Alignment(vertical="center")
        a.border = _border()
        b = ws.cell(row=row, column=2, value=value)
        b.alignment = Alignment(vertical="center", wrap_text=True)
        b.border = _border()

    r = 3
    put(r, "PO Number", po_number);                    r += 1
    put(r, "PO Date", po_date);                        r += 1
    put(r, "Company Name", company);                   r += 1
    put(r, "Buyer Email", email);                      r += 1
    # Ship-To: first line is the name, remaining lines become the address
    put(r, "Ship To Name", ship_to_lines[0]);          r += 1
    put(r, "Ship To", "\n".join(ship_to_lines[1:]));   r += 1
    put(r, "Requested Delivery Date", requested_delivery_date); r += 1

    # ── Line-items table ─────────────────────────────────────────────────────
    r += 1  # blank separator row
    headers = ["Item #", "Product Code", "Description", "Qty", "UOM"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
    ws.row_dimensions[r].height = 22
    r += 1

    for i, item in enumerate(line_items, 1):
        sku, desc, qty = item[0], item[1], item[2]
        uom = item[3] if len(item) > 3 else "EA"
        for ci, val in enumerate([i, sku, desc, qty, uom], 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = Font(size=11)
            cell.fill = PatternFill("solid", fgColor=LBLUE if r % 2 == 0 else "FFFFFF")
            cell.alignment = Alignment(vertical="center", wrap_text=True,
                                       horizontal="center" if ci in (1, 4, 5) else "left")
            cell.border = _border()
        r += 1

    # ── Column widths ────────────────────────────────────────────────────────
    widths = {"A": 22, "B": 44, "C": 38, "D": 10, "E": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    wb.save(path)
    print("Created:", path)


# ── Happy-Flow-PO : clean auto-decision demo (txt twin: Happy-Flow-PO.txt) ────
_write_po(
    os.path.join(OUT_DIR, "Happy-Flow-PO.xlsx"),
    po_number="PO-2026-30001",
    po_date="02 July 2026",
    company="Great Lakes Plumbing Supply Co",
    email="john.miller@glps.com",
    ship_to_lines=[
        "Great Lakes Plumbing - Chicago DC",
        "4500 West Diversey Avenue",
        "Chicago, IL 60639",
    ],
    requested_delivery_date="27 July 2026",
    line_items=[
        ("SKU-CTG-4520", "Ceramic Disc Faucet Cartridge",   100, "EA"),
        ("SKU-SEL-1150", "Tank-to-Bowl Gasket Kit",          120, "EA"),
        ("SKU-VLV-2201", "Pressure-Balancing Shower Valve",   15, "EA"),
    ],
)


# ── CSR-Approval-PO : one file that walks EVERY interactive CSR gate ──────────
#   (txt twin: CSR-Approval-PO.txt). In a single submission it triggers, in
#   order: unresolved buyer, obsolete-SKU substitution, wrong SKU, missing SKU,
#   zero quantity, UOM ambiguity, partial ship-to (which resolves to a remote
#   site) -> then the Pricing & Promo, Inventory and Logistics decision-layer
#   CSR gates. See the CSR demo guide doc for the full walkthrough.
_write_po(
    os.path.join(OUT_DIR, "CSR-Approval-PO.xlsx"),
    po_number="PO-2026-30002",
    po_date="02 July 2026",
    company="Great Lakes Plumbing Supply Co",
    email="procurement@greatlakesps.com",          # not in buyer directory
    ship_to_lines=[
        "Great Lakes - Ketchikan Project Site",     # name only -> partial ship-to
    ],
    requested_delivery_date="28 July 2026",
    line_items=[
        ("SKU-CTG-1000",  "Legacy 2-Handle Faucet Cartridge",      40, "EA"),   # obsolete -> substitute
        ("PN-DRAIN-STD",  "Pop-Up Drain Assembly",                 30, "EA"),   # wrong code -> identify by desc
        ("",              "Tank-to-Bowl Gasket Kit",               50, "EA"),   # missing SKU -> identify by desc
        ("SKU-VLV-2201",  "Pressure-Balancing Shower Valve",        0, "EA"),   # zero qty -> CSR enters qty
        ("SKU-CTG-4520",  "Ceramic Disc Faucet Cartridge",         84, "CASE"), # non-standard UOM -> conversion (1 CASE = 24 EA -> 2016 EA) -> also trips volume-discount pricing exception
        ("SKU-SHS-7700",  "Shower System",                         10, "EA"),   # low stock -> inventory shortage
    ],
)

print("\nAll demo Excel POs generated in demo/")
