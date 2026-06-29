"""
Run this once to generate the sample Excel PO files in sample-data/US-01/.
  python create_sample_excel.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils  import get_column_letter
import os

OUT_DIR = os.path.join(os.path.dirname(__file__), "sample-data", "US-01")
os.makedirs(OUT_DIR, exist_ok=True)

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

thin  = Side(style="thin",   color="AAAAAA")
thick = Side(style="medium", color="1E3A5F")

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def thick_border():
    return Border(left=thick, right=thick, top=thick, bottom=thick)

# ─────────────────────────────────────────────────────────────────────────────
# FILE 1: Happy-path PO (all fields present)
# ─────────────────────────────────────────────────────────────────────────────
def make_happy_path():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Order"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14

    BLUE   = "0369A1"
    LBLUE  = "DBEAFE"
    YELLOW = "FEF9C3"
    GREY   = "F1F5F9"
    WHITE  = "FFFFFF"
    GREEN  = "F0FDF4"

    row = 1

    # Title row
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="PURCHASE ORDER")
    c.font      = Font(bold=True, size=18, color="FFFFFF")
    c.fill      = fill(BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 36
    row += 1

    # ── Header metadata ────────────────────────────────────────────────────────
    header_fields = [
        ("PO Number",               "PO-2026-10010"),
        ("PO Date",                 "24 June 2026"),
        ("Customer ID",             "CUST-1001"),
        ("Company Name",            "Great Lakes Plumbing Supply Co"),
        ("Buyer ID",                "BUY-001"),
        ("Cost Center",             "CC-MW-100"),
        ("Contact Person",          "John Miller"),
        ("Contract Reference",      "CONTRACT-GLP-2026-007"),
        ("Payment Terms",           "Net 30"),
        ("Ship To",                 "Great Lakes Plumbing - Chicago DC, 4500 West Diversey Ave, Chicago IL 60639"),
        ("ZIP",                     "60639"),
        ("Delivery Date",           "15 July 2026"),
        ("Delivery Instructions",   "Deliver to receiving dock. Pallet jack required. Call 312-555-0199 before arrival."),
    ]

    for label, value in header_fields:
        ws.row_dimensions[row].height = 22
        lc = ws.cell(row=row, column=1, value=label)
        lc.font  = Font(bold=True, size=11)
        lc.fill  = fill(GREY)
        lc.alignment = Alignment(vertical="center")
        lc.border = thin_border()

        ws.merge_cells(f"B{row}:G{row}")
        vc = ws.cell(row=row, column=2, value=value)
        vc.font  = Font(size=11)
        vc.fill  = fill(WHITE)
        vc.alignment = Alignment(vertical="center", wrap_text=True)
        vc.border = thin_border()
        row += 1

    row += 1  # blank

    # ── Order lines table ──────────────────────────────────────────────────────
    headers = ["Line", "SKU", "Description", "Quantity", "UOM", "Unit Price", "Line Total"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = Font(bold=True, size=11, color="FFFFFF")
        c.fill      = fill("1E3A5F")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
    ws.row_dimensions[row].height = 26
    row += 1

    lines = [
        (1, "SKU-CTG-4520", "Ceramic Disc Faucet Cartridge",      200, "EA",  10.80,  2160.00),
        (2, "SKU-DRN-3010", "Pop-Up Drain Assembly, Brushed Nickel", 80, "EA", 39.50,  3160.00),
        (3, "SKU-VLV-2201", "Pressure-Balancing Shower Valve",     25, "EA",  80.00,  2000.00),
    ]

    for line_data in lines:
        for ci, val in enumerate(line_data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font  = Font(size=11)
            c.fill  = fill(LBLUE if row % 2 == 0 else WHITE)
            c.alignment = Alignment(horizontal="center" if ci != 3 else "left",
                                     vertical="center")
            c.border = thin_border()
            if ci in (6, 7) and isinstance(val, float):
                c.number_format = '"$"#,##0.00'
        ws.row_dimensions[row].height = 22
        row += 1

    # Total row
    total = round(sum(l[6] for l in lines), 2)
    ws.merge_cells(f"A{row}:F{row}")
    tc = ws.cell(row=row, column=1, value="ORDER TOTAL")
    tc.font = Font(bold=True, size=11); tc.fill=fill(YELLOW)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    tc.border = thin_border()
    vc2 = ws.cell(row=row, column=7, value=total)
    vc2.font = Font(bold=True, size=11); vc2.fill=fill(YELLOW)
    vc2.number_format = '"$"#,##0.00'
    vc2.border = thin_border()
    ws.row_dimensions[row].height = 24

    wb.save(os.path.join(OUT_DIR, "sample-po-happy-path.xlsx"))
    print("Created: sample-po-happy-path.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# FILE 2: PO with missing fields (triggers missing-field exception)
# ─────────────────────────────────────────────────────────────────────────────
def make_missing_fields():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Order"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14

    row = 1
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="PURCHASE ORDER — INCOMPLETE")
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = fill("991B1B")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32
    row += 1

    # Missing: Customer ID, Delivery Date, ZIP
    header_fields = [
        ("PO Number",          "PO-2026-10011"),
        ("PO Date",            "24 June 2026"),
        ("Customer ID",        ""),                    # ← MISSING
        ("Company Name",       "Riverside Mechanical Contractors"),
        ("Contract Reference", ""),                    # ← MISSING
        ("Payment Terms",      ""),                    # ← MISSING
        ("Ship To",            "Riverside Mechanical, Project Site, TX"),
        ("ZIP",                ""),                    # ← MISSING
        ("Delivery Date",      ""),                    # ← MISSING
        ("Delivery Instructions", "Call before delivery"),
    ]
    for label, value in header_fields:
        ws.row_dimensions[row].height = 22
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True, size=11)
        lc.fill = fill("F1F5F9")
        lc.border = thin_border()
        ws.merge_cells(f"B{row}:G{row}")
        vc = ws.cell(row=row, column=2, value=value if value else "— MISSING —")
        vc.font = Font(size=11, color="991B1B" if not value else "000000")
        vc.border = thin_border()
        row += 1

    row += 1
    headers = ["Line", "SKU", "Description", "Quantity", "UOM", "Unit Price", "Line Total"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = fill("1E3A5F")
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
    ws.row_dimensions[row].height = 26
    row += 1

    # One line with missing UOM
    for ci, val in enumerate([1, "SKU-SHS-6601", "Digital Shower Controller", 5, "", 4200.00, 21000.00], 1):
        c = ws.cell(row=row, column=ci, value=val if val != "" else "— MISSING —")
        c.font = Font(size=11, color="991B1B" if val == "" else "000000")
        c.border = thin_border()

    wb.save(os.path.join(OUT_DIR, "sample-po-missing-fields.xlsx"))
    print("Created: sample-po-missing-fields.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# FILE 3: Comprehensive happy-path PO (all possible header + line fields)
# Designed to pass every one of the 12 orchestration stages cleanly.
# ─────────────────────────────────────────────────────────────────────────────
def make_comprehensive():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Order"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16

    BLUE   = "0369A1"
    NAVY   = "1E3A5F"
    LBLUE  = "DBEAFE"
    YELLOW = "FEF9C3"
    GREY   = "F1F5F9"
    WHITE  = "FFFFFF"
    SOFT   = "ECFDF5"

    row = 1

    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="PURCHASE  ORDER")
    c.font      = Font(bold=True, size=18, color="FFFFFF")
    c.fill      = fill(BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 38
    row += 1

    def section(label):
        nonlocal row
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font  = Font(bold=True, size=12, color="FFFFFF")
        c.fill  = fill(NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 26
        row += 1

    def kv(label, value, soft=False):
        nonlocal row
        ws.row_dimensions[row].height = 22
        lc = ws.cell(row=row, column=1, value=label)
        lc.font  = Font(bold=True, size=11)
        lc.fill  = fill(GREY)
        lc.alignment = Alignment(vertical="center")
        lc.border = thin_border()

        ws.merge_cells(f"B{row}:G{row}")
        vc = ws.cell(row=row, column=2, value=value)
        vc.font  = Font(size=11)
        vc.fill  = fill(SOFT if soft else WHITE)
        vc.alignment = Alignment(vertical="center", wrap_text=True)
        vc.border = thin_border()
        row += 1

    # ── PO HEADER ──────────────────────────────────────────────────────────────
    section("PO HEADER")
    kv("PO Number",               "PO-2026-10012")
    kv("PO Date",                 "24 June 2026")
    kv("Document Type",           "Standard B2B Purchase Order")
    kv("Currency",                "USD")
    kv("Priority",                "Normal")

    # ── CUSTOMER / BUYER ──────────────────────────────────────────────────────
    section("CUSTOMER  /  BUYER  DETAILS")
    kv("Customer ID",             "CUST-1001")
    kv("Customer Account Name",   "Great Lakes Plumbing Supply Co")
    kv("ERP Customer ID",         "ERP-GL-7781")
    kv("CRM Account ID",          "CRM-GLP-001")
    kv("Global Parent",           "Continental Building Products Group")
    kv("Regional Division",       "Continental North America")
    kv("Local Branch",            "Great Lakes Midwest Branch (BR-GLP-MW)")
    kv("Buyer ID",                "BUY-001")
    kv("Buyer Name",              "John Miller")
    kv("Buyer Role",              "Senior Buyer")
    kv("Buyer Email",             "john.miller@glps.com")
    kv("Buyer Phone",             "+1-312-555-0192")
    kv("Cost Center",             "CC-MW-100")
    kv("Department",              "Midwest Operations - Procurement")
    kv("Punchout ID",             "PUNCH-GL-001")

    # ── SHIP TO ────────────────────────────────────────────────────────────────
    section("SHIP  TO")
    kv("Ship-To Name",            "Great Lakes Plumbing - Chicago DC")
    kv("Ship-To ID",              "ST-CHI-001")
    kv("Address Line 1",          "4500 West Diversey Avenue")
    kv("Address Line 2",          "Dock 3 - Receiving")
    kv("City",                    "Chicago")
    kv("State",                   "IL")
    kv("Country",                 "USA")
    kv("Ship-To ZIP",             "60639")

    # ── COMMERCIAL TERMS ──────────────────────────────────────────────────────
    section("COMMERCIAL  TERMS")
    kv("Contract Reference",      "CONTRACT-GLP-2026-007")
    kv("Pricing Tier",            "TIER-2")
    kv("Payment Terms",           "Net 30")
    kv("Incoterms",               "DAP (Delivered at Place)")
    kv("Freight Terms",           "Prepaid and Charged")
    kv("Requested Delivery Date", "15 July 2026")
    kv("Required by Site",        "16 July 2026 09:00 CT")
    kv("Preferred Carrier",       "PrimeExpress (Express) / Midwest Freight (Ground)")
    kv("Preferred Warehouse",     "DC-CHI-01 (Chicago Distribution Center)")

    # ── ORDER LINES ───────────────────────────────────────────────────────────
    section("ORDER  LINES")
    headers = ["Line", "SKU", "Description", "Quantity", "UOM", "Unit Price", "Line Total"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = Font(bold=True, size=11, color="FFFFFF")
        c.fill      = fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
    ws.row_dimensions[row].height = 26
    row += 1

    lines = [
        (1, "SKU-CTG-4520", "Ceramic Disc Faucet Cartridge",        250, "EA", 12.50,  3125.00),
        (2, "SKU-DRN-3010", "Pop-Up Drain Assembly, Brushed Nickel", 60, "EA", 45.00,  2700.00),
        (3, "SKU-VLV-2201", "Pressure-Balancing Shower Valve",       20, "EA", 88.00,  1760.00),
        (4, "SKU-SHS-7700", "Digital Shower Interface System",        2, "EA", 1450.00, 2900.00),
        (5, "SKU-SEL-1150", "Tank-to-Bowl Gasket Kit",              200, "EA", 6.25,   1250.00),
    ]

    for line_data in lines:
        for ci, val in enumerate(line_data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font  = Font(size=11)
            c.fill  = fill(LBLUE if row % 2 == 0 else WHITE)
            c.alignment = Alignment(horizontal="center" if ci != 3 else "left",
                                     vertical="center")
            c.border = thin_border()
            if ci in (6, 7) and isinstance(val, (int, float)):
                c.number_format = '"$"#,##0.00'
        ws.row_dimensions[row].height = 22
        row += 1

    total = round(sum(l[6] for l in lines), 2)
    ws.merge_cells(f"A{row}:F{row}")
    tc = ws.cell(row=row, column=1, value="SUB TOTAL (line items at list price)")
    tc.font = Font(bold=True, size=11); tc.fill=fill(YELLOW)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    tc.border = thin_border()
    vc2 = ws.cell(row=row, column=7, value=total)
    vc2.font = Font(bold=True, size=11); vc2.fill=fill(YELLOW)
    vc2.number_format = '"$"#,##0.00'
    vc2.border = thin_border()
    ws.row_dimensions[row].height = 24
    row += 2

    # ── DELIVERY / COMPLIANCE / APPROVALS ─────────────────────────────────────
    section("DELIVERY  &  LOGISTICS  INSTRUCTIONS")
    kv("Delivery Instructions",
       "Deliver to Dock 3. Pallet jack required on site. Call warehouse manager "
       "312-555-0199 before arrival. Provide bill of lading + packing list at gate.")
    kv("Special Instructions",
       "Blanket order against CONTRACT-GLP-2026-007. Partial shipments allowed. "
       "Backorders not required.")

    section("COMPLIANCE  /  DOCUMENTATION")
    kv("Hazardous Materials",     "None on this order")
    kv("SDS Required",            "Not applicable (all SKUs non-hazardous)")
    kv("Country of Origin",       "USA")
    kv("Quality Certificates",    "Standard compliance certificate with shipment")

    section("APPROVALS  &  NOTES")
    kv("Submitted By",            "John Miller (Senior Buyer)", soft=True)
    kv("Authorized",              "Yes - within buyer self-approval limit ($250,000)", soft=True)
    kv("Approval Workflow",       "Auto-approved by buyer authority", soft=True)
    kv("Notes",
       "Standard restock order. Please confirm ETA and tracking once "
       "shipment is dispatched.", soft=True)

    wb.save(os.path.join(OUT_DIR, "sample-po-comprehensive.xlsx"))
    print("Created: sample-po-comprehensive.xlsx")


if __name__ == "__main__":
    make_happy_path()
    make_missing_fields()
    make_comprehensive()
    print("\nAll sample Excel files created in sample-data/US-01/")
