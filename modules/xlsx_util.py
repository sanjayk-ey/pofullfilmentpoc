"""
xlsx_util.py
Shared helpers for reading the mock master-data Excel workbooks.

Every workbook follows the same layout per sheet:
    row 1 = title banner (merged)
    row 2 = column headers
    row 3+ = data rows
"""
import os
import openpyxl

MOCK_DIR = os.path.join(os.path.dirname(__file__), "..", "mock-data")


def mock_path(filename: str) -> str:
    return os.path.join(MOCK_DIR, filename)


def clean(v):
    """Return a stripped string for non-empty values, else None."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def to_num(v, default=None):
    """Best-effort numeric conversion."""
    if v is None or str(v).strip() == "":
        return default
    try:
        f = float(str(v).replace(",", "").replace("$", "").strip())
        return int(f) if f.is_integer() else f
    except (TypeError, ValueError):
        return default


def yes(v) -> bool:
    return str(v).strip().upper() in ("Y", "YES", "TRUE", "1")


def read_sheet(wb, sheet_name: str) -> list:
    """Read a worksheet into a list of row-dicts. Returns [] if the sheet is absent."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[1]]
    out = []
    for raw in rows[2:]:
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        record = {}
        for h, v in zip(headers, raw):
            if h:
                record[h] = v
        out.append(record)
    return out


def load_sheets(filename: str, sheet_names: list) -> dict:
    """Load several sheets from a workbook in one shot: {sheet_name: [rows...]}."""
    path = mock_path(filename)
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        return {name: read_sheet(wb, name) for name in sheet_names}
    finally:
        wb.close()
