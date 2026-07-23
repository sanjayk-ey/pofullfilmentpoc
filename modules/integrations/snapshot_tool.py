"""
snapshot_tool.py
================
Generates the JSON data snapshots that back the mock integration systems
(Mock ERP / PIM / Commerce / OMS / Shipping).

The mock systems serve **exactly** the data that used to be read straight from
the ``mock-data/*.xlsx`` master-data workbooks. This tool reads every workbook
with the SAME row layout used across the app (row 1 = title banner, row 2 =
column headers, row 3+ = data rows) and writes one JSON file per workbook into
``modules/integrations/data/``.

Values are preserved as-is (int / float / str / bool / None). Dates and
datetimes are serialised with ``str(value)`` — the exact string the app already
produces when it renders those cells via ``clean()`` / ``str()`` — so the mock
API returns byte-identical data to the previous Excel-backed path.

Re-run whenever the Excel master data changes:

    python -m modules.integrations.snapshot_tool
"""
import datetime as _dt
import json
import os
from typing import Dict, List

import openpyxl

MOCK_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "mock-data")
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")


def _read_sheet(ws) -> List[dict]:
    """Row 1 = title, row 2 = headers, row 3+ = data. Mirrors xlsx_util.read_sheet
    and account_validator._read_sheet so the snapshot matches load_sheets()."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[1]]
    out = []
    for raw in rows[2:]:
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        record = {}
        for h, v in zip(headers, raw):
            if h:
                record[h] = v
        out.append(record)
    return out


def _jsonable(v):
    """Preserve JSON-native types; render dates/datetimes exactly as the app
    already does (str(value)) so downstream clean()/to_num()/date parsing is
    byte-identical to the Excel-backed path."""
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return str(v)
    return v


def workbook_to_dict(path: str) -> Dict[str, List[dict]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        out = {}
        for name in wb.sheetnames:
            rows = _read_sheet(wb[name])
            out[name] = [{k: _jsonable(v) for k, v in r.items()} for r in rows]
        return out
    finally:
        wb.close()


def generate() -> List[str]:
    os.makedirs(DATA_DIR, exist_ok=True)
    written = []
    for fname in sorted(os.listdir(MOCK_DIR)):
        if not fname.lower().endswith(".xlsx") or fname.startswith("~$"):
            continue
        data = workbook_to_dict(os.path.join(MOCK_DIR, fname))
        out_path = os.path.join(DATA_DIR, fname[:-5] + ".json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=1)
        total = sum(len(v) for v in data.values())
        written.append(out_path)
        print(f"{fname:42s} -> {os.path.basename(out_path):42s} "
              f"({len(data)} sheets, {total} rows)")
    return written


if __name__ == "__main__":
    files = generate()
    print(f"\nGenerated {len(files)} snapshot file(s) in {DATA_DIR}")
