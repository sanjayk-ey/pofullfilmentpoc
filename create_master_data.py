"""
create_master_data.py
Generates ALL domain master-data workbooks used by the orchestration stages:

  mock-data/buyer-master-data.xlsx          (buyer authorization, ordering rights)
  mock-data/product-master-data.xlsx        (product match, variants, UOM, substitution)
  mock-data/compliance-master-data.xlsx     (regional eligibility, restrictions, SDS)
  mock-data/pricing-master-data.xlsx        (price list, contracts, tiers, rebates, surcharges)
  mock-data/budget-master-data.xlsx         (budgets, cost centers, approval matrix, authority)
  mock-data/credit-master-data.xlsx         (credit, invoice aging, payment terms, risk)
  mock-data/inventory-master-data.xlsx      (plant/DC/in-transit stock, ATP, allocation)
  mock-data/logistics-master-data.xlsx      (carrier coverage, freight, SLA, warehouses)
  mock-data/exception-governance-master-data.xlsx (severity, routing, SLA thresholds)
  mock-data/execution-master-data.xlsx      (integration endpoints, comm templates, documents)

Run once:  python create_master_data.py

NOTE: For the POC there is NO real ERP/CRM/WMS/OMS/TMS/SMTP. These workbooks
mock every "Systems / Data Needed" input. The canonical product catalog and the
happy-path customer (CUST-1001 / branch BR-ACME-MW / ship-to ZIP 60639) are kept
consistent across every workbook so the end-to-end pipeline resolves cleanly.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.join(os.path.dirname(__file__), "mock-data")
os.makedirs(OUT_DIR, exist_ok=True)

NAVY = "1E3A5F"
LBLUE = "DBEAFE"
thin = Side(style="thin", color="AAAAAA")


def _border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def write_sheet(ws, title, headers, rows, widths=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[2].height = 30

    for ri, rowdata in enumerate(rows, start=3):
        for ci, val in enumerate(rowdata, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=10)
            cell.fill = PatternFill("solid", fgColor=LBLUE if ri % 2 == 0 else "FFFFFF")
            cell.alignment = Alignment(vertical="center")
            cell.border = _border()

    widths = widths or [16] * len(headers)
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def build_workbook(filename, sheets):
    """sheets = [(sheet_name, title, headers, rows, widths), ...]"""
    wb = openpyxl.Workbook()
    for i, (name, title, headers, rows, widths) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = name
        write_sheet(ws, title, headers, rows, widths)
    out = os.path.join(OUT_DIR, filename)
    wb.save(out)
    print("Created:", out)


# ════════════════════════════════════════════════════════════════════════════
# BUYER MASTER (US-03 — buyer authorization, product visibility, ordering rights)
# ════════════════════════════════════════════════════════════════════════════
build_workbook("buyer-master-data.xlsx", [
    ("Buyer_Profiles",
     "BUYER PROFILES  (buyer identity, role, branch, cost center, authority)",
     ["buyer_id", "buyer_name", "email", "customer_account", "branch_id",
      "default_cost_center", "role", "status", "max_order_value", "currency",
      "can_self_approve", "punchout_id"],
     [
        ["BUY-001", "John Miller", "john.miller@acme.com", "CUST-1001", "BR-ACME-MW", "CC-MW-100", "SENIOR_BUYER", "ACTIVE", 250000, "USD", "Y", "PUNCH-AC-001"],
        ["BUY-002", "Linda Park", "linda.park@acme.com", "CUST-1001", "BR-ACME-MW", "CC-MW-200", "JUNIOR_BUYER", "ACTIVE", 5000, "USD", "N", "PUNCH-AC-002"],
        ["BUY-003", "Mark Snow", "mark.snow@acme.com", "CUST-1002", "BR-ACME-NE", "CC-NE-100", "BUYER", "ACTIVE", 75000, "USD", "Y", "PUNCH-AC-003"],
        ["BUY-010", "Sara Lee", "sara.lee@globex.com", "CUST-5001", "BR-GLOBEX-WEST", "CC-WEST-100", "BUYER", "ACTIVE", 120000, "USD", "Y", "PUNCH-GX-010"],
        ["BUY-900", "Tom Gray (suspended)", "tom.gray@acme.com", "CUST-1001", "BR-ACME-MW", "CC-MW-100", "BUYER", "SUSPENDED", 0, "USD", "N", ""],
     ],
     [12, 20, 26, 16, 16, 16, 14, 12, 14, 8, 14, 14]),

    ("User_Permissions",
     "USER PERMISSIONS  (per-buyer ordering rights and product family access)",
     ["buyer_id", "permitted_branches", "permitted_cost_centers",
      "allowed_product_families", "denied_product_families", "max_line_value",
      "requires_approval_above"],
     [
        ["BUY-001", "BR-ACME-MW", "CC-MW-100,CC-MW-200", "PIPE,FLANGE,VALVE,PUMP,GASKET,CHEMICAL", "", 100000, 250000],
        ["BUY-002", "BR-ACME-MW", "CC-MW-200", "PIPE,FLANGE,GASKET", "PUMP,CHEMICAL,VALVE", 2500, 5000],
        ["BUY-003", "BR-ACME-NE", "CC-NE-100", "PIPE,FLANGE,VALVE,GASKET", "CHEMICAL", 40000, 75000],
        ["BUY-010", "BR-GLOBEX-WEST", "CC-WEST-100", "PIPE,VALVE,PUMP", "CHEMICAL", 60000, 120000],
     ],
     [12, 18, 22, 34, 24, 14, 18]),

    ("Cost_Centers",
     "COST CENTERS  (validity, owning branch, status)",
     ["cost_center_id", "name", "branch_id", "status", "owner", "currency"],
     [
        ["CC-MW-100", "Midwest Operations", "BR-ACME-MW", "ACTIVE", "John Miller", "USD"],
        ["CC-MW-200", "Midwest Maintenance", "BR-ACME-MW", "ACTIVE", "Linda Park", "USD"],
        ["CC-NE-100", "Northeast Operations", "BR-ACME-NE", "ACTIVE", "Mark Snow", "USD"],
        ["CC-WEST-100", "West Operations", "BR-GLOBEX-WEST", "ACTIVE", "Sara Lee", "USD"],
        ["CC-OLD-900", "Decommissioned Center", "BR-ACME-MW", "INACTIVE", "-", "USD"],
     ],
     [16, 24, 16, 10, 16, 8]),

    ("Product_Visibility_Rules",
     "PRODUCT VISIBILITY RULES  (catalog visibility by hierarchy/cost center)",
     ["rule_id", "scope_type", "scope_id", "product_family", "sku",
      "visibility", "min_role", "reason"],
     [
        ["PV-001", "global_parent", "GP-ACME", "ALL", "", "VISIBLE", "JUNIOR_BUYER", "Default catalog"],
        ["PV-002", "product_family", "CHEMICAL", "CHEMICAL", "", "RESTRICTED", "SENIOR_BUYER", "Hazardous goods - senior buyers only"],
        ["PV-003", "sku", "", "PUMP", "SKU-PMP-7700", "RESTRICTED", "SENIOR_BUYER", "Capital equipment approval"],
        ["PV-004", "cost_center", "CC-MW-200", "VALVE", "", "HIDDEN", "BUYER", "Valves not orderable for maintenance CC"],
     ],
     [10, 16, 16, 16, 16, 14, 16, 30]),
])


# ════════════════════════════════════════════════════════════════════════════
# PRODUCT MASTER (US-04 — matching, configuration, variant, UOM)
# ════════════════════════════════════════════════════════════════════════════
build_workbook("product-master-data.xlsx", [
    ("Product_Master",
     "PRODUCT MASTER  (variants, status, base UOM, material, grade, configuration)",
     ["sku", "description", "product_family", "status", "base_uom", "material",
      "grade", "size", "configurable", "hazardous", "substitute_sku",
      "list_price", "currency", "weight_kg", "lead_time_days", "manufacturer",
      "country_of_origin"],
     [
        ["SKU-STL-4520", "Carbon Steel Pipe 4 inch SCH40", "PIPE", "ACTIVE", "FT", "Carbon Steel", "A106-B", "4 inch", "Y", "N", "", 12.50, "USD", 16.0, 7, "SteelCorp", "USA"],
        ["SKU-FLG-3010", "Stainless Steel Flange 3 inch 150#", "FLANGE", "ACTIVE", "EA", "SS316", "150#", "3 inch", "Y", "N", "", 45.00, "USD", 5.5, 10, "FlangeTech", "USA"],
        ["SKU-VLV-2201", "Ball Valve 2 inch Full Port", "VALVE", "ACTIVE", "EA", "SS304", "Class 600", "2 inch", "Y", "N", "", 88.00, "USD", 3.2, 14, "ValveWorks", "Germany"],
        ["SKU-PMP-7700", "Centrifugal Pump 5HP", "PUMP", "ACTIVE", "EA", "Cast Iron", "API-610", "5HP", "Y", "N", "", 1450.00, "USD", 65.0, 28, "PumpPro", "USA"],
        ["SKU-CHM-9100", "Industrial Solvent Drum", "CHEMICAL", "ACTIVE", "GAL", "Solvent", "Tech-Grade", "55 gal", "N", "Y", "", 9.75, "USD", 200.0, 5, "ChemSafe", "USA"],
        ["SKU-CHM-9200", "Aerosol Degreaser (no SDS on file)", "CHEMICAL", "ACTIVE", "GAL", "Solvent", "Tech-Grade", "20 gal", "N", "Y", "", 14.00, "USD", 80.0, 6, "ChemSafe", "USA"],
        ["SKU-GSK-1150", "Spiral Wound Gasket 3 inch", "GASKET", "ACTIVE", "EA", "Graphite/SS", "Class 150", "3 inch", "N", "N", "", 6.25, "USD", 0.3, 4, "SealTech", "USA"],
        ["SKU-OBS-1000", "Legacy Carbon Pipe 4 inch (obsolete)", "PIPE", "OBSOLETE", "FT", "Carbon Steel", "A53", "4 inch", "N", "N", "SKU-STL-4520", 11.00, "USD", 15.0, 0, "SteelCorp", "USA"],
        ["SKU-INA-2000", "Discontinued Gate Valve 2 inch", "VALVE", "INACTIVE", "EA", "Cast Steel", "Class 300", "2 inch", "N", "N", "SKU-VLV-2201", 70.00, "USD", 4.0, 0, "ValveWorks", "Germany"],
     ],
     [14, 32, 14, 10, 9, 14, 10, 8, 12, 10, 14, 11, 8, 10, 13, 12, 14]),

    ("Product_Attributes",
     "PRODUCT ATTRIBUTES  (required configuration attributes per variant)",
     ["sku", "attribute_name", "attribute_value", "required", "uom_dimension"],
     [
        ["SKU-STL-4520", "schedule", "SCH40", "Y", ""],
        ["SKU-STL-4520", "coating", "Black", "N", ""],
        ["SKU-FLG-3010", "pressure_class", "150#", "Y", ""],
        ["SKU-FLG-3010", "face_type", "Raised Face", "Y", ""],
        ["SKU-VLV-2201", "connection_type", "Flanged", "Y", ""],
        ["SKU-VLV-2201", "actuation", "Manual", "N", ""],
        ["SKU-PMP-7700", "voltage", "460V", "Y", ""],
        ["SKU-PMP-7700", "seal_type", "Mechanical", "Y", ""],
        ["SKU-CHM-9100", "concentration", "99%", "Y", ""],
     ],
     [14, 18, 16, 10, 14]),

    ("UOM_Conversions",
     "UOM CONVERSIONS  (approved conversion factors)",
     ["from_uom", "to_uom", "factor", "product_family", "notes"],
     [
        ["M", "FT", 3.28084, "PIPE", "metres to feet"],
        ["FT", "M", 0.3048, "PIPE", "feet to metres"],
        ["L", "GAL", 0.264172, "CHEMICAL", "litres to US gallons"],
        ["GAL", "L", 3.78541, "CHEMICAL", "US gallons to litres"],
        ["BOX", "EA", 12, "GASKET", "1 box = 12 each"],
        ["PALLET", "EA", 50, "FLANGE", "1 pallet = 50 each"],
        ["KG", "LB", 2.20462, "ALL", "kilograms to pounds"],
     ],
     [10, 10, 12, 16, 22]),

    ("Substitution_Rules",
     "SUBSTITUTION RULES  (approved substitutes for obsolete / inactive SKUs)",
     ["original_sku", "substitute_sku", "compatibility", "price_impact_pct",
      "availability_impact", "requires_approval", "rationale"],
     [
        ["SKU-OBS-1000", "SKU-STL-4520", "FULL", 13.6, "IN_STOCK", "Y", "A106-B supersedes A53; same dimensions"],
        ["SKU-INA-2000", "SKU-VLV-2201", "FUNCTIONAL", 25.7, "IN_STOCK", "Y", "Ball valve replaces discontinued gate valve"],
     ],
     [14, 14, 14, 14, 16, 16, 38]),

    ("Compatibility_Rules",
     "COMPATIBILITY RULES  (cross-product compatibility constraints)",
     ["sku", "compatible_with_sku", "rule_type", "notes"],
     [
        ["SKU-FLG-3010", "SKU-GSK-1150", "RECOMMENDED", "3 inch flange pairs with 3 inch gasket"],
        ["SKU-VLV-2201", "SKU-FLG-3010", "RECOMMENDED", "Flanged valve requires matching flanges"],
        ["SKU-PMP-7700", "SKU-VLV-2201", "OPTIONAL", "Isolation valve recommended on pump suction"],
     ],
     [14, 18, 14, 36]),
])


# ════════════════════════════════════════════════════════════════════════════
# COMPLIANCE MASTER (US-05 — regional eligibility, restrictions, SDS)
# ════════════════════════════════════════════════════════════════════════════
build_workbook("compliance-master-data.xlsx", [
    ("Compliance_Rules",
     "COMPLIANCE RULES  (product-region restrictions, approvals, authority)",
     ["rule_id", "sku", "product_family", "restriction_type", "region_state",
      "status", "reason", "approval_required", "authority"],
     [
        ["CMP-001", "SKU-CHM-9100", "CHEMICAL", "REGIONAL", "CA", "RESTRICTED", "VOC limits in California", "Y", "EPA/CARB"],
        ["CMP-002", "SKU-CHM-9100", "CHEMICAL", "REGIONAL", "NY", "ALLOWED", "Permitted with SDS", "N", "EPA"],
        ["CMP-003", "", "CHEMICAL", "EXPORT", "UK", "PROHIBITED", "Import licence required - not held", "Y", "HSE"],
        ["CMP-004", "SKU-PMP-7700", "PUMP", "REGIONAL", "ALL", "ALLOWED", "No restriction", "N", "-"],
     ],
     [10, 14, 14, 14, 12, 12, 30, 16, 12]),

    ("Regional_Restrictions",
     "REGIONAL RESTRICTIONS  (region/state -> restricted families)",
     ["region_code", "region_name", "restricted_families", "notes"],
     [
        ["IL", "Illinois", "", "No regional product restrictions"],
        ["MI", "Michigan", "", "No regional product restrictions"],
        ["NY", "New York", "", "Chemicals allowed with SDS attached"],
        ["CA", "California", "CHEMICAL", "VOC-restricted chemicals require approval"],
        ["UK", "United Kingdom", "CHEMICAL", "Export licence required for chemicals"],
     ],
     [12, 16, 20, 36]),

    ("SDS_Repository",
     "SAFETY DATA SHEET REPOSITORY  (required documents per product)",
     ["sku", "sds_document_id", "sds_version", "issue_date", "expiry_date",
      "hazard_class", "required", "file_ref"],
     [
        ["SKU-CHM-9100", "SDS-CHM-9100", "v4.2", "2025-02-01", "2027-02-01", "Class 3 Flammable", "Y", "sds/SKU-CHM-9100_v4.2.pdf"],
     ],
     [14, 16, 12, 14, 14, 18, 10, 28]),

    ("Product_Eligibility",
     "PRODUCT ELIGIBILITY  (per region sale/ship eligibility)",
     ["sku", "region", "eligible", "conditions"],
     [
        ["SKU-STL-4520", "ALL", "Y", "None"],
        ["SKU-FLG-3010", "ALL", "Y", "None"],
        ["SKU-VLV-2201", "ALL", "Y", "None"],
        ["SKU-PMP-7700", "ALL", "Y", "None"],
        ["SKU-GSK-1150", "ALL", "Y", "None"],
        ["SKU-CHM-9100", "IL", "Y", "SDS must be attached"],
        ["SKU-CHM-9100", "CA", "N", "Restricted - requires compliance approval"],
        ["SKU-CHM-9100", "UK", "N", "Export prohibited without licence"],
     ],
     [14, 10, 10, 34]),
])


# ════════════════════════════════════════════════════════════════════════════
# PRICING MASTER (US-06 — multi-layer enterprise B2B pricing)
# ════════════════════════════════════════════════════════════════════════════
build_workbook("pricing-master-data.xlsx", [
    ("Price_List",
     "PRICE LIST  (base list price per SKU)",
     ["sku", "currency", "list_price", "uom", "effective_from", "effective_to", "price_list_id"],
     [
        ["SKU-STL-4520", "USD", 12.50, "FT", "2026-01-01", "2026-12-31", "PL-2026"],
        ["SKU-FLG-3010", "USD", 45.00, "EA", "2026-01-01", "2026-12-31", "PL-2026"],
        ["SKU-VLV-2201", "USD", 88.00, "EA", "2026-01-01", "2026-12-31", "PL-2026"],
        ["SKU-PMP-7700", "USD", 1450.00, "EA", "2026-01-01", "2026-12-31", "PL-2026"],
        ["SKU-CHM-9100", "USD", 9.75, "GAL", "2026-01-01", "2026-12-31", "PL-2026"],
        ["SKU-CHM-9200", "USD", 14.00, "GAL", "2026-01-01", "2026-12-31", "PL-2026"],
        ["SKU-GSK-1150", "USD", 6.25, "EA", "2026-01-01", "2026-12-31", "PL-2026"],
     ],
     [14, 8, 12, 8, 14, 14, 12]),

    ("Contracts",
     "CONTRACTS  (negotiated, date-bound customer pricing)",
     ["contract_reference", "customer_account", "scope_type", "scope_id",
      "contract_price", "currency", "valid_from", "valid_to", "status"],
     [
        ["CONTRACT-ACME-2026-007", "CUST-1001", "sku", "SKU-STL-4520", 10.80, "USD", "2026-01-01", "2026-12-31", "ACTIVE"],
        ["CONTRACT-ACME-2026-007", "CUST-1001", "sku", "SKU-FLG-3010", 39.50, "USD", "2026-01-01", "2026-12-31", "ACTIVE"],
        ["CONTRACT-ACME-2026-007", "CUST-1001", "family", "VALVE", 80.00, "USD", "2026-01-01", "2026-12-31", "ACTIVE"],
        ["CONTRACT-ACME-2025-099", "CUST-1001", "sku", "SKU-PMP-7700", 1300.00, "USD", "2025-01-01", "2025-12-31", "EXPIRED"],
        ["CONTRACT-GLOBEX-2026-001", "CUST-5001", "family", "PIPE", 11.90, "USD", "2026-01-01", "2026-12-31", "ACTIVE"],
     ],
     [24, 16, 12, 16, 14, 8, 12, 12, 10]),

    ("Volume_Tiers",
     "VOLUME TIERS  (tiered discounts by quantity)",
     ["tier_id", "scope_type", "scope_id", "min_qty", "max_qty", "discount_pct", "uom"],
     [
        ["VT-PIPE-1", "family", "PIPE", 0, 499, 0, "FT"],
        ["VT-PIPE-2", "family", "PIPE", 500, 1999, 5, "FT"],
        ["VT-PIPE-3", "family", "PIPE", 2000, 999999, 10, "FT"],
        ["VT-FLG-1", "family", "FLANGE", 0, 99, 0, "EA"],
        ["VT-FLG-2", "family", "FLANGE", 100, 999, 4, "EA"],
        ["VT-VLV-1", "family", "VALVE", 0, 49, 0, "EA"],
        ["VT-VLV-2", "family", "VALVE", 50, 999, 6, "EA"],
     ],
     [12, 12, 14, 10, 10, 12, 8]),

    ("Rebates",
     "REBATES  (customer rebate accruals)",
     ["rebate_id", "customer_account", "product_family", "rebate_pct", "condition"],
     [
        ["RB-ACME-01", "CUST-1001", "PIPE", 2.0, "Annual volume > 10000 FT"],
        ["RB-ACME-02", "CUST-1001", "ALL", 1.0, "Loyalty rebate"],
     ],
     [12, 16, 14, 12, 28]),

    ("Promotions",
     "PROMOTIONS  (time-bound promotional discounts)",
     ["promo_id", "scope_type", "scope_id", "discount_pct", "valid_from", "valid_to", "description"],
     [
        ["PROMO-Q3-GASKET", "family", "GASKET", 8.0, "2026-06-01", "2026-09-30", "Q3 gasket promotion"],
     ],
     [16, 12, 14, 12, 14, 14, 26]),

    ("Surcharges",
     "SURCHARGES  (location/regional fees)",
     ["surcharge_id", "scope_type", "scope_id", "surcharge_type", "amount_type", "amount", "reason"],
     [
        ["SUR-LA-001", "zip", "90001", "REGIONAL_FEE", "PCT", 1.5, "West coast handling surcharge"],
        ["SUR-FUEL-01", "ALL", "ALL", "FUEL", "PCT", 0.75, "Fuel adjustment"],
     ],
     [14, 12, 14, 16, 12, 10, 28]),

    ("Freight_Terms",
     "FREIGHT TERMS  (incoterms and freight responsibility)",
     ["scope_type", "scope_id", "incoterm", "freight_payer", "base_freight", "notes"],
     [
        ["customer", "CUST-1001", "FOB Destination", "SELLER", 150.00, "Prepaid & add"],
        ["customer", "CUST-5001", "FOB Origin", "BUYER", 0.00, "Collect"],
        ["default", "ALL", "FOB Destination", "SELLER", 200.00, "Default freight"],
     ],
     [12, 16, 16, 14, 12, 18]),

    ("Margin_Policy",
     "MARGIN POLICY  (minimum margin / max discount per family)",
     ["product_family", "min_margin_pct", "max_discount_pct", "approver_role"],
     [
        ["PIPE", 15, 18, "PRICING_MANAGER"],
        ["FLANGE", 18, 15, "PRICING_MANAGER"],
        ["VALVE", 20, 12, "PRICING_MANAGER"],
        ["PUMP", 22, 10, "PRICING_DIRECTOR"],
        ["GASKET", 25, 20, "PRICING_MANAGER"],
        ["CHEMICAL", 12, 8, "PRICING_MANAGER"],
     ],
     [16, 14, 16, 18]),

    ("Raw_Material_Index",
     "RAW MATERIAL INDEX  (material cost adjustments)",
     ["material", "index_date", "index_pct_adjustment", "notes"],
     [
        ["Carbon Steel", "2026-06-01", 1.5, "Steel index up 1.5%"],
        ["SS316", "2026-06-01", 2.2, "Stainless surcharge"],
        ["Cast Iron", "2026-06-01", 0.0, "No change"],
     ],
     [16, 14, 18, 24]),
])


# ════════════════════════════════════════════════════════════════════════════
# BUDGET MASTER (US-07 — budget, spend limit, approval routing)
# ════════════════════════════════════════════════════════════════════════════
build_workbook("budget-master-data.xlsx", [
    ("Budget_Master",
     "BUDGET MASTER  (available budget by hierarchy level / period)",
     ["level_type", "level_id", "period", "budget_amount", "consumed_amount",
      "available_amount", "currency"],
     [
        ["global_parent", "GP-ACME", "2026", 50000000, 18000000, 32000000, "USD"],
        ["regional_division", "RD-ACME-NA", "2026", 12000000, 5000000, 7000000, "USD"],
        ["branch", "BR-ACME-MW", "2026", 2000000, 1200000, 800000, "USD"],
        ["branch", "BR-ACME-NE", "2026", 1500000, 1400000, 100000, "USD"],
        ["branch", "BR-GLOBEX-WEST", "2026", 1800000, 600000, 1200000, "USD"],
     ],
     [16, 16, 10, 14, 14, 14, 8]),

    ("Cost_Centers",
     "COST CENTER BUDGETS  (cost-center level spend control)",
     ["cost_center_id", "name", "branch_id", "status", "budget_amount",
      "consumed_amount", "available_amount", "currency"],
     [
        ["CC-MW-100", "Midwest Operations", "BR-ACME-MW", "ACTIVE", 600000, 250000, 350000, "USD"],
        ["CC-MW-200", "Midwest Maintenance", "BR-ACME-MW", "ACTIVE", 80000, 78000, 2000, "USD"],
        ["CC-NE-100", "Northeast Operations", "BR-ACME-NE", "ACTIVE", 300000, 295000, 5000, "USD"],
        ["CC-WEST-100", "West Operations", "BR-GLOBEX-WEST", "ACTIVE", 400000, 100000, 300000, "USD"],
     ],
     [16, 22, 16, 10, 14, 14, 14, 8]),

    ("Approval_Matrix",
     "APPROVAL MATRIX  (spend thresholds -> approver role per level)",
     ["level_type", "level_id", "min_amount", "max_amount", "approver_role",
      "approver_name", "sla_hours"],
     [
        ["branch", "BR-ACME-MW", 0, 50000, "BRANCH_MANAGER", "Karen Wells", 24],
        ["regional_division", "RD-ACME-NA", 50001, 250000, "REGIONAL_MANAGER", "David Cho", 48],
        ["global_parent", "GP-ACME", 250001, 99999999, "CORPORATE_PROCUREMENT", "Procurement Board", 72],
        ["branch", "BR-GLOBEX-WEST", 0, 60000, "BRANCH_MANAGER", "Ana Reyes", 24],
     ],
     [16, 16, 12, 14, 20, 16, 10]),

    ("Buyer_Authority",
     "BUYER AUTHORITY  (self-approval limits)",
     ["buyer_id", "max_order_value", "can_self_approve", "branch_id", "cost_centers"],
     [
        ["BUY-001", 250000, "Y", "BR-ACME-MW", "CC-MW-100,CC-MW-200"],
        ["BUY-002", 5000, "N", "BR-ACME-MW", "CC-MW-200"],
        ["BUY-003", 75000, "Y", "BR-ACME-NE", "CC-NE-100"],
        ["BUY-010", 120000, "Y", "BR-GLOBEX-WEST", "CC-WEST-100"],
     ],
     [12, 16, 16, 16, 22]),
])


# ════════════════════════════════════════════════════════════════════════════
# CREDIT MASTER (US-08 — credit, payment terms, financial risk)
# ════════════════════════════════════════════════════════════════════════════
build_workbook("credit-master-data.xlsx", [
    ("Credit_Master",
     "CREDIT MASTER  (limit, available credit, status, risk)",
     ["customer_account", "credit_limit", "available_credit", "currency",
      "payment_terms", "credit_status", "risk_rating"],
     [
        ["CUST-1001", 1000000, 650000, "USD", "NET30", "GOOD", "LOW"],
        ["CUST-1002", 500000, 20000, "USD", "NET45", "WATCH", "MEDIUM"],
        ["CUST-2001", 400000, 250000, "USD", "NET30", "GOOD", "LOW"],
        ["CUST-5001", 800000, 500000, "USD", "NET60", "GOOD", "LOW"],
        ["CUST-7000", 200000, 0, "USD", "PREPAYMENT", "HOLD", "HIGH"],
     ],
     [16, 14, 16, 8, 14, 12, 10]),

    ("Invoice_Aging",
     "INVOICE AGING  (open AR and overdue invoices)",
     ["customer_account", "invoice_no", "amount", "due_date", "days_overdue", "status"],
     [
        ["CUST-1001", "INV-AC-5501", 120000, "2026-06-30", 0, "OPEN"],
        ["CUST-1002", "INV-AC-6602", 95000, "2026-05-15", 42, "OVERDUE"],
        ["CUST-1002", "INV-AC-6610", 60000, "2026-05-30", 27, "OVERDUE"],
        ["CUST-7000", "INV-IN-9901", 200000, "2026-04-01", 86, "OVERDUE"],
     ],
     [16, 14, 12, 14, 14, 10]),

    ("Payment_History",
     "PAYMENT HISTORY  (payment behaviour)",
     ["customer_account", "avg_days_to_pay", "last_payment_date", "last_payment_amount", "dispute_count"],
     [
        ["CUST-1001", 28, "2026-06-01", 130000, 0],
        ["CUST-1002", 51, "2026-05-01", 40000, 2],
        ["CUST-2001", 30, "2026-06-10", 75000, 0],
        ["CUST-5001", 55, "2026-06-05", 200000, 0],
        ["CUST-7000", 95, "2026-02-15", 50000, 4],
     ],
     [16, 16, 16, 18, 14]),

    ("Payment_Terms",
     "PAYMENT TERMS  (terms catalog)",
     ["terms_code", "description", "net_days", "prepayment_required"],
     [
        ["NET30", "Net 30 days", 30, "N"],
        ["NET45", "Net 45 days", 45, "N"],
        ["NET60", "Net 60 days", 60, "N"],
        ["PREPAYMENT", "Prepayment required", 0, "Y"],
     ],
     [14, 22, 10, 18]),

    ("Risk_Signals",
     "RISK SIGNALS  (fraud / watchlist indicators)",
     ["customer_account", "fraud_flag", "risk_score", "watchlist", "notes"],
     [
        ["CUST-1001", "N", 12, "N", "Stable account"],
        ["CUST-1002", "N", 55, "N", "Slow payer"],
        ["CUST-7000", "Y", 88, "Y", "Multiple disputes, on credit hold"],
     ],
     [16, 12, 12, 12, 28]),
])


# ════════════════════════════════════════════════════════════════════════════
# INVENTORY MASTER (US-09 — availability, ATP, allocation, partial fulfilment)
# ════════════════════════════════════════════════════════════════════════════
build_workbook("inventory-master-data.xlsx", [
    ("Plant_Stock",
     "PLANT STOCK  (manufacturing plant on-hand)",
     ["location_id", "location_type", "region", "sku", "on_hand_qty", "uom"],
     [
        ["PLANT-CHI", "PLANT", "Midwest", "SKU-STL-4520", 5000, "FT"],
        ["PLANT-CHI", "PLANT", "Midwest", "SKU-FLG-3010", 800, "EA"],
        ["PLANT-CHI", "PLANT", "Midwest", "SKU-VLV-2201", 60, "EA"],
        ["PLANT-CHI", "PLANT", "Midwest", "SKU-PMP-7700", 4, "EA"],
        ["PLANT-CHI", "PLANT", "Midwest", "SKU-GSK-1150", 1500, "EA"],
        ["PLANT-CHI", "PLANT", "Midwest", "SKU-CHM-9100", 900, "GAL"],
     ],
     [12, 14, 12, 14, 12, 8]),

    ("DC_Stock",
     "DISTRIBUTION CENTER STOCK  (DC on-hand by region)",
     ["location_id", "location_type", "region", "sku", "on_hand_qty", "uom"],
     [
        ["DC-CHI-01", "DC", "Midwest", "SKU-STL-4520", 3000, "FT"],
        ["DC-CHI-01", "DC", "Midwest", "SKU-FLG-3010", 400, "EA"],
        ["DC-CHI-01", "DC", "Midwest", "SKU-VLV-2201", 30, "EA"],
        ["DC-CHI-01", "DC", "Midwest", "SKU-GSK-1150", 600, "EA"],
        ["DC-CHI-01", "DC", "Midwest", "SKU-PMP-7700", 8, "EA"],
        ["DC-DET-02", "DC", "Midwest", "SKU-STL-4520", 1200, "FT"],
        ["DC-DET-02", "DC", "Midwest", "SKU-VLV-2201", 15, "EA"],
        ["DC-LA-05", "DC", "West", "SKU-STL-4520", 2000, "FT"],
     ],
     [12, 14, 12, 14, 12, 8]),

    ("In_Transit",
     "IN-TRANSIT INVENTORY  (inbound stock with ETA)",
     ["sku", "from_location", "to_location", "qty", "uom", "eta_date"],
     [
        ["SKU-VLV-2201", "PLANT-CHI", "DC-CHI-01", 40, "EA", "2026-07-05"],
        ["SKU-PMP-7700", "PLANT-CHI", "DC-CHI-01", 3, "EA", "2026-07-12"],
     ],
     [14, 14, 14, 10, 8, 14]),

    ("ATP",
     "AVAILABLE-TO-PROMISE  (net ATP and replenishment)",
     ["sku", "atp_qty", "uom", "next_replenishment_date", "replenishment_qty"],
     [
        ["SKU-STL-4520", 8000, "FT", "2026-07-15", 5000],
        ["SKU-FLG-3010", 1100, "EA", "2026-07-20", 500],
        ["SKU-VLV-2201", 90, "EA", "2026-07-05", 40],
        ["SKU-PMP-7700", 5, "EA", "2026-07-12", 3],
        ["SKU-GSK-1150", 2100, "EA", "2026-07-10", 1000],
        ["SKU-CHM-9100", 900, "GAL", "2026-07-08", 500],
     ],
     [14, 10, 8, 18, 16]),

    ("Allocation_Rules",
     "ALLOCATION RULES  (priority when inventory is constrained)",
     ["customer_tier", "priority", "backorder_tolerance_days", "notes"],
     [
        ["PLATINUM", 1, 14, "Highest allocation priority"],
        ["GOLD", 2, 10, "Standard contract priority"],
        ["SILVER", 3, 7, "Best effort"],
     ],
     [14, 10, 22, 28]),

    ("Fulfillment_Preferences",
     "FULFILLMENT PREFERENCES  (per-customer fulfilment rules)",
     ["customer_account", "customer_tier", "preferred_warehouse", "restricted_dc",
      "split_shipment", "backorder_tolerance_days"],
     [
        ["CUST-1001", "GOLD", "DC-CHI-01", "", "ALLOW", 10],
        ["CUST-1002", "SILVER", "DC-CHI-01", "DC-LA-05", "DENY", 5],
        ["CUST-5001", "GOLD", "DC-LA-05", "", "ALLOW", 10],
     ],
     [16, 14, 18, 14, 14, 22]),
])


# ════════════════════════════════════════════════════════════════════════════
# LOGISTICS MASTER (US-10 — serviceability, SLA, freight, optimization)
# ════════════════════════════════════════════════════════════════════════════
build_workbook("logistics-master-data.xlsx", [
    ("Carrier_Coverage",
     "CARRIER COVERAGE  (serviceable ZIP prefixes per carrier)",
     ["carrier", "zip_prefix", "serviceable", "service_level", "transit_days"],
     [
        ["FastFreight", "606", "Y", "GROUND", 2],
        ["FastFreight", "482", "Y", "GROUND", 3],
        ["FastFreight", "100", "Y", "GROUND", 3],
        ["FastFreight", "900", "Y", "GROUND", 4],
        ["ExpressCo", "606", "Y", "EXPRESS", 1],
        ["ExpressCo", "100", "Y", "EXPRESS", 1],
        ["RegionalCarrier", "999", "N", "GROUND", 0],
     ],
     [16, 10, 12, 14, 12]),

    ("Freight_Rating",
     "FREIGHT RATING  (rate by carrier, zone, weight band)",
     ["carrier", "zone", "weight_min", "weight_max", "base_rate", "per_kg_rate"],
     [
        ["FastFreight", "Z1", 0, 100, 50, 0.80],
        ["FastFreight", "Z2", 0, 100, 75, 1.10],
        ["FastFreight", "Z3", 0, 100, 110, 1.50],
        ["ExpressCo", "Z1", 0, 100, 120, 2.00],
     ],
     [16, 8, 12, 12, 12, 12]),

    ("SLA_Rules",
     "SLA RULES  (delivery commitments per customer tier)",
     ["scope_type", "scope_id", "service_level", "max_transit_days", "on_time_target_pct"],
     [
        ["customer", "CUST-1001", "GROUND", 3, 98],
        ["customer", "CUST-5001", "GROUND", 4, 95],
        ["default", "ALL", "GROUND", 5, 95],
     ],
     [12, 16, 14, 16, 18]),

    ("Warehouse_Master",
     "WAREHOUSE MASTER  (shipping warehouses)",
     ["warehouse_id", "name", "zip", "region", "capacity", "cutoff_time", "carriers"],
     [
        ["DC-CHI-01", "Chicago DC", "60639", "Midwest", "HIGH", "16:00", "FastFreight,ExpressCo"],
        ["DC-DET-02", "Detroit DC", "48201", "Midwest", "MEDIUM", "15:00", "FastFreight"],
        ["DC-LA-05", "Los Angeles DC", "90001", "West", "HIGH", "17:00", "FastFreight"],
     ],
     [14, 16, 10, 12, 10, 12, 22]),

    ("Delivery_Calendar",
     "DELIVERY CALENDAR  (ship days and cutoffs per warehouse)",
     ["warehouse_id", "weekday", "cutoff_time", "ships", "holiday_dates"],
     [
        ["DC-CHI-01", "Mon-Fri", "16:00", "Y", "2026-07-04,2026-12-25"],
        ["DC-DET-02", "Mon-Fri", "15:00", "Y", "2026-07-04,2026-12-25"],
        ["DC-LA-05", "Mon-Sat", "17:00", "Y", "2026-12-25"],
     ],
     [14, 12, 12, 8, 24]),
])


# ════════════════════════════════════════════════════════════════════════════
# EXCEPTION GOVERNANCE MASTER (US-11 — severity, routing, SLA)
# ════════════════════════════════════════════════════════════════════════════
build_workbook("exception-governance-master-data.xlsx", [
    ("Severity_Matrix",
     "SEVERITY MATRIX  (exception type -> severity, category, owning role)",
     ["exception_type", "severity", "category", "default_role"],
     [
        ["MISSING_FIELDS", "MEDIUM", "INTAKE", "CSR"],
        ["DUPLICATE_PO", "MEDIUM", "INTAKE", "CSR"],
        ["UNMATCHED_CUSTOMER", "HIGH", "CUSTOMER", "CSR"],
        ["DUPLICATE_CUSTOMER", "HIGH", "CUSTOMER", "CSR"],
        ["INVALID_SHIP_TO", "MEDIUM", "CUSTOMER", "CSR"],
        ["HIERARCHY_MISMATCH", "HIGH", "CUSTOMER", "ACCOUNT_MANAGER"],
        ["UNAUTHORIZED_BUYER", "HIGH", "AUTHORIZATION", "ACCOUNT_MANAGER"],
        ["RESTRICTED_PRODUCT", "MEDIUM", "AUTHORIZATION", "ACCOUNT_MANAGER"],
        ["INVALID_COST_CENTER", "MEDIUM", "AUTHORIZATION", "CSR"],
        ["PRODUCT_CONFIG_EXCEPTION", "MEDIUM", "PRODUCT", "PRODUCT_SPECIALIST"],
        ["OBSOLETE_SKU", "LOW", "PRODUCT", "PRODUCT_SPECIALIST"],
        ["INVALID_UOM", "MEDIUM", "PRODUCT", "PRODUCT_SPECIALIST"],
        ["COMPLIANCE_RESTRICTION", "HIGH", "COMPLIANCE", "COMPLIANCE_APPROVER"],
        ["MISSING_SDS", "MEDIUM", "COMPLIANCE", "COMPLIANCE_APPROVER"],
        ["PRICING_EXCEPTION", "HIGH", "PRICING", "PRICING_APPROVER"],
        ["EXPIRED_CONTRACT", "MEDIUM", "PRICING", "PRICING_APPROVER"],
        ["BUDGET_EXCEEDED", "HIGH", "APPROVAL", "APPROVER"],
        ["APPROVAL_REQUIRED", "MEDIUM", "APPROVAL", "APPROVER"],
        ["CREDIT_HOLD", "HIGH", "CREDIT", "FINANCE"],
        ["INVENTORY_SHORTAGE", "MEDIUM", "INVENTORY", "FULFILLMENT_PLANNER"],
        ["ALLOCATION_CONFLICT", "MEDIUM", "INVENTORY", "FULFILLMENT_PLANNER"],
        ["ZIP_NOT_SERVICEABLE", "HIGH", "LOGISTICS", "LOGISTICS"],
        ["SLA_MISS", "MEDIUM", "LOGISTICS", "LOGISTICS"],
        ["EXECUTION_FAILURE", "HIGH", "EXECUTION", "CSR"],
     ],
     [26, 10, 16, 22]),

    ("Role_Routing",
     "ROLE ROUTING  (category -> handling role, escalation, queue)",
     ["category", "role", "escalation_role", "queue"],
     [
        ["INTAKE", "CSR", "SUPERVISOR", "Q-INTAKE"],
        ["CUSTOMER", "CSR", "ACCOUNT_MANAGER", "Q-CUSTOMER"],
        ["AUTHORIZATION", "ACCOUNT_MANAGER", "SUPERVISOR", "Q-AUTH"],
        ["PRODUCT", "PRODUCT_SPECIALIST", "SUPERVISOR", "Q-PRODUCT"],
        ["COMPLIANCE", "COMPLIANCE_APPROVER", "COMPLIANCE_HEAD", "Q-COMPLIANCE"],
        ["PRICING", "PRICING_APPROVER", "PRICING_DIRECTOR", "Q-PRICING"],
        ["APPROVAL", "APPROVER", "REGIONAL_MANAGER", "Q-APPROVAL"],
        ["CREDIT", "FINANCE", "FINANCE_MANAGER", "Q-CREDIT"],
        ["INVENTORY", "FULFILLMENT_PLANNER", "SUPERVISOR", "Q-INVENTORY"],
        ["LOGISTICS", "LOGISTICS", "LOGISTICS_MANAGER", "Q-LOGISTICS"],
        ["EXECUTION", "CSR", "IT_SUPPORT", "Q-EXECUTION"],
     ],
     [16, 20, 20, 14]),

    ("SLA_Thresholds",
     "SLA THRESHOLDS  (resolution and escalation timers by severity)",
     ["severity", "sla_hours", "escalation_hours", "notify_channel"],
     [
        ["HIGH", 4, 8, "EMAIL+SMS"],
        ["MEDIUM", 24, 48, "EMAIL"],
        ["LOW", 72, 120, "EMAIL"],
     ],
     [10, 12, 18, 16]),
])


# ════════════════════════════════════════════════════════════════════════════
# EXECUTION MASTER (US-12 — downstream creation, communication, audit)
# ════════════════════════════════════════════════════════════════════════════
build_workbook("execution-master-data.xlsx", [
    ("Integration_Endpoints",
     "INTEGRATION ENDPOINTS  (mock downstream systems)",
     ["system", "record_type", "endpoint", "mode", "default_status", "retry_limit"],
     [
        ["ERP", "Sales Order", "mock://erp/salesorder", "MOCK", "SUCCESS", 3],
        ["OMS", "Order Request", "mock://oms/order", "MOCK", "SUCCESS", 3],
        ["WMS", "Pick Ticket", "mock://wms/pick", "MOCK", "SUCCESS", 3],
        ["TMS", "Shipment Order", "mock://tms/shipment", "MOCK", "SUCCESS", 3],
        ["SMTP", "Confirmation Email", "mock://smtp/send", "MOCK", "SUCCESS", 2],
     ],
     [10, 16, 26, 8, 14, 10]),

    ("Communication_Templates",
     "COMMUNICATION TEMPLATES  (customer messaging)",
     ["template_id", "channel", "subject", "body_ref"],
     [
        ["TPL-ORDER-CONF", "EMAIL", "Your order {order_no} is confirmed", "templates/order_confirmation.txt"],
        ["TPL-EXCEPTION", "EMAIL", "Action needed on order {order_no}", "templates/exception_notice.txt"],
     ],
     [16, 10, 34, 30]),

    ("Document_Repository",
     "DOCUMENT REPOSITORY  (attachments per order/product)",
     ["doc_id", "doc_type", "linked_to", "file_ref"],
     [
        ["DOC-SDS-CHM", "SDS", "SKU-CHM-9100", "sds/SKU-CHM-9100_v4.2.pdf"],
        ["DOC-TC-2026", "Terms", "ALL", "docs/terms_2026.pdf"],
     ],
     [14, 12, 16, 30]),
])

print("\nAll master-data workbooks created in mock-data/")
